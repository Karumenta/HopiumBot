import discord
from discord.ext import commands, tasks
import logging
from dotenv import load_dotenv
import os
import requests
import asyncio
import aiohttp
from aiohttp import web
import threading
import json
import time
import csv
import requests
import os
import re
import logging
import zipfile

from pathlib import Path

from datetime import datetime, timedelta

import json
import time
import csv
import requests
import os

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from datetime import datetime, timedelta

load_dotenv()
token = os.getenv('DISCORD_TOKEN')

# Configure logging for both local and Render environments
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),  # This outputs to stdout/stderr for Render logs
        logging.FileHandler('hopiumbot.log', encoding='utf-8', mode='a') if not os.getenv('RENDER') else logging.NullHandler()
    ]
)

# Store ongoing applications
active_applications = {}

# Application timeout settings (in seconds)
PATH_SELECTION_TIMEOUT = 1200  # 20 minutes for path selection
QUESTION_WARNING_TIME = 600   # 10 minutes - send warning
QUESTION_TIMEOUT = 900         # 15 minutes - cancel application

# Set discord.py logging level to reduce spam
logging.getLogger('discord').setLevel(logging.WARNING)
logging.getLogger('discord.http').setLevel(logging.WARNING)

# Create logger for our bot
logger = logging.getLogger('HopiumBot')
intents = discord.Intents.default()
intents.message_content = True  # Enable message content intent
intents.guilds = True  # Enable guild intents
intents.members = True  # Enable member intents

# Application questions
# Branching Application System
APPLICATION_CONFIG = {
    "intro": {
        "question": "Welcome to <Hopium>! As we have two raiding groups, please indicate which one you are interested in:",
        "options": [
            {"id": "speed_run", "label": "⚡ Speed Run", "description": "Monday 20:00-00:00 ST - Alt character preferred"},
            {"id": "chill", "label": "😎 Chill", "description": "Thursday 20:00-00:00 ST - 1 Character required"},
            {"id": "both", "label": "🎯 Both", "description": "Both raiding days - Alt character preferred for Speed Run"},
            {"id": "none", "label": "❌ None", "description": "I'm not interested"}
        ]
    },
    "paths": {
        "speed_run": [
            "[Speedrun] Main character name:",
            "[Speedrun] Main Class/Spec:",
            "[Speedrun] Alt character name:",
            "[Speedrun] Alt Class/Spec:",
            "What country are you from and how old are you?",
            "Please tell us a bit about yourself, who are you outside of the game?",
            "Explain your WoW experience. Include logs of past relevant characters (Classic/SoM//SoD/Retail).",
            "We require a few things from every raider in the guild. To have above average performance for your class and atleast 80% raid attendance. Can you fulfill these requirements?",
            "Why did you choose to apply to <Hopium>?",
            "Can someone in <Hopium> vouch for you?",
            "Surprise us! What's something you'd like to tell us, it can be absolutely anything!"
        ],
        "chill": [
            "[Chill run] Character name:",
            "[Chill run] Class/Spec:",
            "What country are you from and how old are you?",
            "Please tell us a bit about yourself, who are you outside of the game?",
            "Explain your WoW experience. Include logs of past relevant characters (Classic/SoM//SoD/Retail).",
            "We require a few things from every raider in the guild. To have above average performance for your class and atleast 80% raid attendance. Can you fulfill these requirements?",
            "Why did you choose to apply to <Hopium>?",
            "Can someone in <Hopium> vouch for you?",
            "Surprise us! What's something you'd like to tell us, it can be absolutely anything!"
        ],
        "both": [
            "[Speedrun] Main character name:",
            "[Speedrun] Main Class/Spec:",
            "[Speedrun] Alt character name:",
            "[Speedrun] Alt Class/Spec:",
            "[Chill run] Character name:",
            "[Chill run] Class/Spec:",
            "What country are you from and how old are you?",
            "Please tell us a bit about yourself, who are you outside of the game?",
            "Explain your WoW experience. Include logs of past relevant characters (Classic/SoM//SoD/Retail).",
            "We require a few things from every raider in the guild. To have above average performance for your class and atleast 80% raid attendance. Can you fulfill these requirements?",
            "Why did you choose to apply to <Hopium>?",
            "Can someone in <Hopium> vouch for you?",
            "Surprise us! What's something you'd like to tell us, it can be absolutely anything!"
        ]
    }
}

# Question indices (0-based) that are optional per path.
# When the user types 'skip' at any question in the set, all questions in that set are skipped.
OPTIONAL_QUESTIONS = {
    "speed_run": {2, 3},  # Alt character name and Alt Class/Spec
    "both": {2, 3},       # Alt character name and Alt Class/Spec (Speedrun alt)
}

# Legacy support - will be removed after migration
APPLICATION_QUESTIONS = APPLICATION_CONFIG["paths"]["both"]

CLASS_LIST = {
    "Druid" : {"name": "Druid", "roles": ["DPS", "Heal", "Tank"], "color": "FF7C0A"},
    "Hunter" : {"name": "Hunter", "roles": ["DPS"], "color": "AAD372"},
    "Mage" : {"name": "Mage", "roles": ["DPS"], "color": "3FC7EB"},
    "Paladin" : {"name": "Paladin", "roles": ["DPS", "Heal", "Tank"], "color": "F48CBA"},
    "Priest" : {"name": "Priest", "roles": ["Heal", "DPS"], "color": "FFFFFF"},
    "Rogue" : {"name": "Rogue", "roles": ["DPS"], "color": "FFF468"},
    "Shaman" : {"name": "Shaman", "roles": ["DPS", "Heal"], "color": "0070DD"},
    "Warlock" : {"name": "Warlock", "roles": ["DPS"], "color": "8788EE"},
    "Warrior" : {"name": "Warrior", "roles": ["DPS", "Tank"], "color": "C69B6D"}
}

BLIZZARD_ID = os.getenv('BLIZZARD_ID')
BLIZZARD_SECRET = os.getenv('BLIZZARD_SECRET')
BLIZZARD_TOKEN_URL = 'https://eu.battle.net/oauth/token'

WCL_ID = os.getenv('WCL_ID')
WCL_SECRET = os.getenv('WCL_SECRET')

# Validate required environment variables
missing_vars = []
if not BLIZZARD_ID:
    missing_vars.append('BLIZZARD_ID')
if not BLIZZARD_SECRET:
    missing_vars.append('BLIZZARD_SECRET')
if not WCL_ID:
    missing_vars.append('WCL_ID')
if not WCL_SECRET:
    missing_vars.append('WCL_SECRET')

if missing_vars:
    error_msg = f"Missing required environment variables: {', '.join(missing_vars)}"
    print(f"❌ {error_msg}")
    logger.error(error_msg)
    print("Please set these variables in your .env file or Render environment variables.")
else:
    logger.info("All required API credentials loaded successfully")

bot = commands.Bot(command_prefix='!', intents=intents)

role = "Trial"

# Define guild-specific paths that work both locally and on Render
def get_guild_data_path(guild_id):
    """Get data path specific to a guild"""
    if os.getenv('RENDER'):
        # Production on Render
        base_path = '/app/data'
    else:
        # Local development
        base_dir = os.path.dirname(os.path.abspath(__file__))
        base_path = os.path.join(base_dir, 'app', 'data')
    
    # Create guild-specific directory
    guild_path = os.path.join(base_path, f'guild_{guild_id}')
    os.makedirs(guild_path, exist_ok=True)
    return guild_path

def get_guild_file_paths(guild_id):
    """Get all file paths for a specific guild"""
    guild_data_dir = get_guild_data_path(guild_id)
    
    # TMB directory and files
    tmb_dir = os.path.join(guild_data_dir, 'tmb')
    character_file = os.path.join(tmb_dir, 'character-json.json')
    attendance_file = os.path.join(tmb_dir, 'hopium-attendance.csv')
    item_file = os.path.join(tmb_dir, 'item-notes.csv')
    
    # Cache directory and files
    cache_dir = os.path.join(guild_data_dir, 'cache')
    armory_file = os.path.join(cache_dir, 'armory.json')
    item_icons_file = os.path.join(cache_dir, 'item-icons.json')
    parses_file = os.path.join(cache_dir, 'parses.json')
    
    # Sheet directory
    sheet_dir = os.path.join(guild_data_dir, 'sheets')
    
    return {
        'guild_data_dir': guild_data_dir,
        'tmb_dir': tmb_dir,
        'cache_dir': cache_dir,
        'sheet_dir': sheet_dir,
        'character_file': character_file,
        'attendance_file': attendance_file,
        'item_file': item_file,
        'armory_file': armory_file,
        'item_icons_file': item_icons_file,
        'parses_file': parses_file
    }

def calculate_gradient_color(value, start_color, end_color):
    value = max(0, min(1, value))

    start_red, start_green, start_blue = start_color
    end_red, end_green, end_blue = end_color

    red = int(start_red + (end_red - start_red) * value)
    green = int(start_green + (end_green - start_green) * value)
    blue = int(start_blue + (end_blue - start_blue) * value)

    return f"{red:02X}{green:02X}{blue:02X}"

# Initialize guild-specific data files
def initialize_guild_data_files(guild_id):
    """Initialize required directories and files for a specific guild"""
    try:
        paths = get_guild_file_paths(guild_id)
        
        # Create directories
        os.makedirs(paths['tmb_dir'], exist_ok=True)
        os.makedirs(paths['cache_dir'], exist_ok=True)
        os.makedirs(paths['sheet_dir'], exist_ok=True)
        
        # Initialize character file if it doesn't exist
        if not os.path.exists(paths['character_file']):
            with open(paths['character_file'], 'w', encoding='utf-8') as f:
                json.dump([], f)
            print(f"📄 Created empty character file for guild {guild_id}")
        
        # Initialize armory file if it doesn't exist
        if not os.path.exists(paths['armory_file']):
            with open(paths['armory_file'], 'w', encoding='utf-8') as f:
                json.dump({}, f)
            print(f"📄 Created empty armory file for guild {guild_id}")
        
        # Initialize item icons file if it doesn't exist
        if not os.path.exists(paths['item_icons_file']):
            with open(paths['item_icons_file'], 'w', encoding='utf-8') as f:
                json.dump({}, f)
            print(f"📄 Created empty item icons file for guild {guild_id}")
        
        # Initialize parses file if it doesn't exist
        if not os.path.exists(paths['parses_file']):
            with open(paths['parses_file'], 'w', encoding='utf-8') as f:
                json.dump({}, f)
            print(f"📄 Created empty parses file for guild {guild_id}")
        
        print(f"✅ Data files initialized successfully for guild {guild_id}")
        logger.info(f"Data files initialized successfully for guild {guild_id}")
        
    except Exception as e:
        print(f"❌ Error initializing data files for guild {guild_id}: {e}")
        logger.error(f"Error initializing data files for guild {guild_id}: {e}", exc_info=True)

# Application cleanup functions
async def check_application_timeouts():
    """Check for applications that need warnings or cleanup due to inactivity"""
    current_time = asyncio.get_event_loop().time()
    warnings_sent = []
    cancelled_apps = []
    
    print(f"🔍 Checking {len(active_applications)} active applications for timeouts")
    
    for user_id, app_data in list(active_applications.items()):
        # Skip if user is still in path selection phase (handled by Discord UI timeout)
        if app_data.get('question_index', -1) < 0:
            print(f"  Skipping user {user_id}: still in path selection phase")
            continue
            
        last_activity = app_data.get('last_activity', app_data.get('start_time', current_time))
        inactive_time = current_time - last_activity
        question_index = app_data.get('question_index', 0)
        warning_sent = app_data.get('warning_sent', False)
        
        print(f"  User {user_id}: inactive for {inactive_time:.1f}s, question {question_index}, warning_sent: {warning_sent}")
        
        # Check if user needs a warning (10 minutes of inactivity)
        if inactive_time >= QUESTION_WARNING_TIME and not warning_sent:
            print(f"  🚨 Sending warning to user {user_id}")
            try:
                user = bot.get_user(user_id)
                if user:
                    questions = app_data.get('questions', [])
                    current_q = app_data.get('question_index', 0)
                    
                    embed = discord.Embed(
                        title="⚠️ Inactivity Detected",
                        description="You have **5 minutes** to answer the current question or your application will be automatically cancelled.",
                        color=0xff6b00
                    )
                    
                    if current_q < len(questions):
                        embed.add_field(
                            name=f"Current Question ({current_q + 1}/{len(questions)})",
                            value=questions[current_q],
                            inline=False
                        )
                    
                    await user.send(embed=embed)
                    app_data['warning_sent'] = True
                    warnings_sent.append(user_id)
                    
            except Exception as e:
                logger.warning(f"Could not send warning to user {user_id}: {e}")
        
        # Check if application should be cancelled (15 minutes total inactivity)
        elif inactive_time >= QUESTION_TIMEOUT:
            print(f"  ❌ Cancelling application for user {user_id}")
            try:
                user = bot.get_user(user_id)
                if user:
                    embed = discord.Embed(
                        title="❌ Application Cancelled",
                        description="Your application has been cancelled due to inactivity (no response for 15 minutes).",
                        color=0xff0000
                    )
                    embed.add_field(
                        name="Want to reapply?",
                        value="You can start a new application anytime by clicking the Apply button again in the server.",
                        inline=False
                    )
                    await user.send(embed=embed)
                    
                cancelled_apps.append(user_id)
                del active_applications[user_id]
                
            except Exception as e:
                logger.warning(f"Could not notify user {user_id} about cancellation: {e}")
                cancelled_apps.append(user_id)
                if user_id in active_applications:
                    del active_applications[user_id]
    
    if warnings_sent:
        print(f"📤 Sent inactivity warnings to {len(warnings_sent)} users")
        logger.info(f"Sent inactivity warnings to {len(warnings_sent)} users")
    if cancelled_apps:
        print(f"🧹 Cleaned up {len(cancelled_apps)} inactive applications")
        logger.info(f"Cancelled {len(cancelled_apps)} inactive applications")
    if not warnings_sent and not cancelled_apps:
        print("✅ No timeout actions needed")

# Separate task for frequent timeout checking
@tasks.loop(seconds=30)  # Check every 30 seconds for timeouts
async def timeout_checker_task():
    try:
        if active_applications:  # Only run if there are active applications
            await check_application_timeouts()
    except Exception as e:
        print(f"❌ Error in timeout checker: {e}")
        logger.error(f"Error in timeout checker: {e}", exc_info=True)

@timeout_checker_task.before_loop
async def before_timeout_checker():
    await bot.wait_until_ready()
    print("⏰ Timeout checker started - will run every 30 seconds")
    logger.info("Timeout checker started - will run every 30 seconds")

# Background task that runs every X minutes
@tasks.loop(minutes=5)  # Run every 5 minutes to properly handle application timeouts
async def periodic_task():
    try:
        current_time = datetime.now().strftime('%H:%M:%S')
        print(f"🔄 Starting periodic update for all guilds at {current_time}")
        logger.info(f"Starting periodic update for all guilds at {current_time}")
        
        # Only run the full data update every cycle (no timeout checking here - handled by separate task)
        print(f"📊 Running full data update cycle")
        logger.info(f"Running full data update cycle")
        
        # Process each guild the bot is in
        for guild in bot.guilds:
                try:
                    print(f"🏰 Processing guild: {guild.name} (ID: {guild.id})")
                    logger.info(f"Processing guild: {guild.name} (ID: {guild.id})")
                    
                    # Initialize guild data files if needed
                    initialize_guild_data_files(guild.id)
                    
                    # Get guild-specific file paths
                    paths = get_guild_file_paths(guild.id)
                    
                    # Check if required files exist for this guild
                    if not os.path.exists(paths['character_file']):
                        print(f"⚠️ Character file not found for guild {guild.name}: {paths['character_file']}")
                        continue
                    
                    # Process this guild's data
                    await process_guild_data(guild.id, paths)
                    
                except Exception as e:
                    print(f"❌ Error processing guild {guild.name}: {e}")
                    logger.error(f"Error processing guild {guild.name}: {e}", exc_info=True)
                    continue
        
        print(f"✅ Periodic data update completed for all guilds")
        logger.info(f"Periodic data update completed for all guilds")
        
    except Exception as e:
        print(f"❌ Critical error in periodic task: {e}")
        logger.error(f"Critical error in periodic task: {e}", exc_info=True)

async def process_guild_data(guild_id, paths):
    """Process data for a specific guild"""
    # Load players with better error handling
    players = {}
    try:
        with open(paths['character_file'], 'r', encoding='utf-8') as file:
            character_data = json.load(file)
            for playerInfo in character_data:
                player_name = playerInfo.get('name', '').strip()
                if player_name:
                    # Handle display_archetype - default to "DPS" if None or empty
                    display_archetype = playerInfo.get('display_archetype')
                    if display_archetype is None or not display_archetype.strip():
                        archetype = "DPS"
                    else:
                        archetype = display_archetype.strip()
                    
                    player = {
                        "name": player_name.capitalize(),
                        "archetype": archetype,
                    }
                    players[player_name.lower()] = player
        
        if not players:
            print("ℹ️ No players found in character file")
            return
            
        print(f"📋 Processing {len(players)} characters")
    except (json.JSONDecodeError, FileNotFoundError) as e:
        print(f"❌ Error loading character file: {e}")
        return
    
    try:
        # Check if required API credentials are available
        if not BLIZZARD_ID or not BLIZZARD_SECRET:
            error_msg = "Blizzard API credentials not available. Skipping armory update."
            print(f"❌ {error_msg}")
            logger.error(error_msg)
            return
        
        logger.info(f"Using Blizzard API credentials - ID: {BLIZZARD_ID[:8]}..., Secret: {'*' * len(BLIZZARD_SECRET)}")
        
        # Get Blizzard API token with retry logic
        access_token = None
        for attempt in range(3):
            try:
                logger.info(f"Attempting to get Blizzard API token (attempt {attempt + 1}/3)")
                response = requests.post(
                    BLIZZARD_TOKEN_URL, 
                    data={'grant_type': 'client_credentials'}, 
                    auth=(BLIZZARD_ID, BLIZZARD_SECRET),
                    timeout=10
                )
                response.raise_for_status()
                access_token = response.json()['access_token']
                logger.info("Successfully obtained Blizzard API token")
                break
            except requests.RequestException as e:
                error_msg = f"Token request attempt {attempt + 1} failed: {e}"
                print(f"⚠️ {error_msg}")
                logger.warning(error_msg)
                if attempt == 2:
                    final_error = "Failed to get Blizzard API token after 3 attempts"
                    print(f"❌ {final_error}")
                    logger.error(final_error)
                    return
                await asyncio.sleep(2)  # Wait before retry
        
        # Load existing armory data using guild-specific paths
        armory_data = {}
        if os.path.exists(paths['armory_file']):
            try:
                with open(paths['armory_file'], "r", encoding="utf-8") as f:
                    armory_data = json.load(f)
            except (json.JSONDecodeError, FileNotFoundError):
                print("⚠️ Creating new armory file")
                armory_data = {}
        
        # Load existing parses data using guild-specific paths
        parses_data = {}
        if os.path.exists(paths['parses_file']):
            try:
                with open(paths['parses_file'], "r", encoding="utf-8") as f:
                    parses_data = json.load(f)
            except (json.JSONDecodeError, FileNotFoundError):
                print("⚠️ Creating new parses file")
                parses_data = {}
        
        # Get WCL API token with retry logic
        wcl_access_token = None
        if WCL_ID and WCL_SECRET:
            for attempt in range(3):
                try:
                    logger.info(f"Attempting to get WCL API token (attempt {attempt + 1}/3)")
                    wcl_token_url = "https://fresh.warcraftlogs.com/oauth/token"
                    wcl_data = {"grant_type": "client_credentials"}
                    wcl_response = requests.post(wcl_token_url, data=wcl_data, auth=(WCL_ID, WCL_SECRET), timeout=10)
                    wcl_response.raise_for_status()
                    wcl_access_token = wcl_response.json()["access_token"]
                    logger.info("Successfully obtained WCL API token")
                    break
                except requests.RequestException as e:
                    error_msg = f"WCL token request attempt {attempt + 1} failed: {e}"
                    print(f"⚠️ {error_msg}")
                    logger.warning(error_msg)
                    if attempt == 2:
                        final_error = "Failed to get WCL API token after 3 attempts"
                        print(f"❌ {final_error}")
                        logger.error(final_error)
                    await asyncio.sleep(2)  # Wait before retry
        else:
            logger.warning("WCL API credentials not available. Skipping WCL data updates.")
        
        # Fetch equipment and parses for each character with rate limiting
        new_items_found = 0
        new_parses_found = 0
        characters_processed = 0
        
        for player_key, player_info in players.items():
            try:
                # Rate limiting - small delay between API calls
                if characters_processed > 0:
                    await asyncio.sleep(0.7)  # 700ms delay between calls for both APIs
                
                player_name = player_info["name"]
                character_name = player_name.lower().replace(" ", "-")
                
                # Fetch WCL parses data (if WCL token is available)
                if wcl_access_token:
                    try:
                        wcl_server_slug = "spineshatter"
                        wcl_server_region = "EU"
                        wcl_api_url = "https://fresh.warcraftlogs.com/api/v2/client"
                        wcl_headers = {"Authorization": f"Bearer {wcl_access_token}"}
                        
                        # Determine metric based on archetype
                        metric = "hps" if player_info.get("archetype", "").lower() == "healer" else "bossdps"
                        
                        query = f"""
                        {{
                            characterData {{
                                character(name: "{player_key.lower()}", serverSlug: "{wcl_server_slug}", serverRegion: "{wcl_server_region}") {{
                                    zoneRankings(metric: {metric})
                                }}
                            }}
                        }}
                        """
                        
                        response = requests.post(wcl_api_url, headers=wcl_headers, json={"query": query}, timeout=10)
                        if response.status_code == 200:
                            wcl_data = response.json()
                            try:
                                # Extract zone rankings
                                character_data = wcl_data.get("data", {}).get("characterData", {})
                                character_info = character_data.get("character", {}) if character_data else {}
                                rankings = character_info.get("zoneRankings", {}) if character_info else {}
                                
                                # Parse rankings if it's a string
                                if isinstance(rankings, str):
                                    rankings = json.loads(rankings)
                                
                                # Create structured parse data
                                parse_info = {
                                    "metric": metric,
                                    "archetype": player_info.get("archetype", ""),
                                    "bestPerformanceAverage": rankings.get("bestPerformanceAverage", 0.0),
                                    "medianPerformanceAverage": rankings.get("medianPerformanceAverage", 0.0),
                                    "lastUpdated": datetime.now().isoformat()
                                }
                                
                                # Check if parses data changed
                                existing_parses = parses_data.get(player_name, {})
                                if (existing_parses.get("bestPerformanceAverage") != parse_info["bestPerformanceAverage"] or
                                    existing_parses.get("medianPerformanceAverage") != parse_info["medianPerformanceAverage"]):
                                    
                                    parses_data[player_name] = parse_info
                                    new_parses_found += 1
                                    print(f"📊 {player_name}: Updated parses (Best: {parse_info['bestPerformanceAverage']:.1f}, Median: {parse_info['medianPerformanceAverage']:.1f})")
                                
                            except (KeyError, json.JSONDecodeError, TypeError) as e:
                                print(f"⚠️ Error parsing WCL data for {player_name}: {e}")
                                # Set default values if parsing fails
                                parses_data[player_name] = {
                                    "metric": metric,
                                    "archetype": player_info.get("archetype", ""),
                                    "bestPerformanceAverage": 0.0,
                                    "medianPerformanceAverage": 0.0,
                                    "lastUpdated": datetime.now().isoformat(),
                                    "error": "Failed to parse WCL data"
                                }
                        elif response.status_code == 404:
                            print(f"⚠️ Character not found on WCL: {player_name}")
                        elif response.status_code == 429:
                            print(f"⚠️ WCL rate limited for {player_name}, waiting...")
                            await asyncio.sleep(5)
                        else:
                            print(f"⚠️ WCL API error for {player_name}: {response.status_code}")
                            
                    except Exception as e:
                        print(f"⚠️ Error fetching WCL parses for {player_name}: {str(e)[:100]}")
                
                # Fetch Blizzard armory data
                try:
                    url = f"https://eu.api.blizzard.com/profile/wow/character/spineshatter/{character_name}/equipment"
                    params = {
                        "namespace": "profile-classicann-eu",
                        "locale": "en_GB"
                    }
                    headers = {'Authorization': f'Bearer {access_token}'}
                    
                    logger.debug(f"Fetching armory data for {player_name} from {url}")
                    
                    # Use async HTTP client for better performance
                    async with aiohttp.ClientSession() as session:
                        async with session.get(url, params=params, headers=headers, timeout=10) as response:
                            if response.status == 200:
                                data = await response.json()
                                equipped_items = [item["name"] for item in data.get("equipped_items", [])]
                                
                                # Initialize player in armory_data if not exists
                                if player_name not in armory_data:
                                    armory_data[player_name] = []
                                
                                # Check for new items
                                existing_items = set(armory_data[player_name])
                                new_items = [item for item in equipped_items if item not in existing_items]
                                
                                if new_items:
                                    armory_data[player_name].extend(new_items)
                                    new_items_found += len(new_items)
                                    print(f"🆕 {player_name}: {len(new_items)} new items")
                                    logger.info(f"Found {len(new_items)} new items for {player_name}")
                                    
                                    # Log each new item
                                    for item in new_items:
                                        logger.info(f"New item found for {player_name}: {item}")
                                else:
                                    logger.debug(f"No new items for {player_name}")
                            
                            elif response.status == 401:
                                error_msg = f"Unauthorized request for {player_name} - token may be expired"
                                print(f"❌ {error_msg}")
                                logger.error(error_msg)
                            elif response.status == 403:
                                error_msg = f"Forbidden request for {player_name} - check API credentials"
                                print(f"❌ {error_msg}")
                                logger.error(error_msg)
                            elif response.status == 404:
                                print(f"⚠️ Character not found on Blizzard API: {player_name}")
                                logger.warning(f"Character not found on Blizzard API: {player_name}")
                            elif response.status == 429:
                                print(f"⚠️ Blizzard API rate limited for {player_name}, waiting...")
                                logger.warning(f"Blizzard API rate limited for {player_name}")
                                await asyncio.sleep(5)
                            else:
                                error_msg = f"Blizzard API error for {player_name}: HTTP {response.status}"
                                print(f"⚠️ {error_msg}")
                                logger.error(error_msg)
                                # Log response text for debugging
                                try:
                                    response_text = await response.text()
                                    logger.error(f"Response body: {response_text[:200]}")
                                except:
                                    pass
                
                except aiohttp.ClientError as e:
                    error_msg = f"Network error fetching Blizzard armory for {player_name}: {str(e)}"
                    print(f"⚠️ {error_msg}")
                    logger.error(error_msg)
                except Exception as e:
                    error_msg = f"Unexpected error fetching Blizzard armory for {player_name}: {str(e)}"
                    print(f"⚠️ {error_msg}")
                    logger.error(error_msg, exc_info=True)
                
                characters_processed += 1
                
            except asyncio.TimeoutError:
                print(f"⚠️ Timeout fetching data for {player_name}")
            except Exception as e:
                print(f"⚠️ Error processing {player_name}: {str(e)[:100]}")
        
        # Save updated armory data atomically
        if new_items_found > 0 or characters_processed > 0:
            try:
                # Sort armory data alphabetically by character name before saving
                sorted_armory_data = dict(sorted(armory_data.items()))
                
                # Write to temporary file first, then rename (atomic operation)
                temp_file = paths['armory_file'] + '.tmp'
                with open(temp_file, "w", encoding="utf-8") as f:
                    json.dump(sorted_armory_data, f, ensure_ascii=False, indent=2)
                
                # Atomic rename
                os.replace(temp_file, paths['armory_file'])
                print(f"💾 Armory data saved - {new_items_found} new items found")
            except Exception as e:
                print(f"❌ Error saving armory data: {e}")
                # Clean up temp file if it exists
                if os.path.exists(temp_file):
                    os.remove(temp_file)
        
        # Save updated parses data atomically
        if new_parses_found > 0 or characters_processed > 0:
            try:
                # Write to temporary file first, then rename (atomic operation)
                temp_file = paths['parses_file'] + '.tmp'
                with open(temp_file, "w", encoding="utf-8") as f:
                    json.dump(parses_data, f, ensure_ascii=False, indent=2)
                
                # Atomic rename
                os.replace(temp_file, paths['parses_file'])
                print(f"� Parses data saved - {new_parses_found} characters updated")
            except Exception as e:
                print(f"❌ Error saving parses data: {e}")
                # Clean up temp file if it exists
                if os.path.exists(temp_file):
                    os.remove(temp_file)
        
        print(f"✅ Data update completed - {characters_processed}/{len(players)} characters processed")
        logger.info(f"Data update completed - {characters_processed}/{len(players)} characters processed")
        print(f"   📊 Summary: {new_items_found} new items, {new_parses_found} parse updates")
        logger.info(f"Summary: {new_items_found} new items, {new_parses_found} parse updates")
        
    except Exception as e:
        print(f"❌ Critical error in periodic task: {e}")
        logger.error(f"Critical error in periodic task: {e}", exc_info=True)

@periodic_task.before_loop
async def before_periodic_task():
    await bot.wait_until_ready()
    print("🚀 Data update task started - will run every 5 minutes")
    logger.info("Data update task started - will run every 5 minutes")

async def validate_character_exists(character_name):
    try:
        url = f"https://classicwowarmory.com/character/eu/spineshatter/{character_name.replace(' ', '%20')}"
        
        # Check if the character exists
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=10) as response:
                if "/character/404" in str(response.url) or response.status != 200:
                    return False, "Character not found on armory"
                return True, None
                    
    except Exception as e:
        return False, f"Error checking character: {str(e)}"

async def get_staff_mentions(guild):
    mentions = []
    
    # Look for Karumenta
    karumenta = discord.utils.get(guild.members, name="Karumenta")
    if karumenta:
        mentions.append(karumenta.mention)
    else:
        mentions.append("**Karumenta**")
    
    # Look for Hokkies
    hokkies = discord.utils.get(guild.members, name="Hokkies")
    if hokkies:
        mentions.append(hokkies.mention)
    else:
        mentions.append("**Hokkies**")
    
    return " or ".join(mentions)

async def validate_character_name(character_name, guild=None):
    try:
        url = f"https://classicwowarmory.com/character/eu/spineshatter/{character_name.replace(' ', '%20')}"
        response = requests.get(url, timeout=10, allow_redirects=True)
        print(url)
        
        # Get staff mentions
        staff_mentions = "**Karumenta** or **Hokkies**"
        if guild:
            staff_mentions = await get_staff_mentions(guild)
        
        # Check if we were redirected to the 404 page
        if "/character/404" in response.url or response.status_code == 404:
            return False, f"Character not found on Spineshatter realm. Please check the spelling and try again. If it's an error, please contact {staff_mentions} for assistance."
        elif response.status_code == 200:
            return True, None
        else:
            return False, f"Unable to verify character (HTTP {response.status_code}). Please contact {staff_mentions} for assistance."
    
    except requests.exceptions.Timeout:
        staff_mentions = "**Karumenta** or **Hokkies**"
        if guild:
            staff_mentions = await get_staff_mentions(guild)
        return False, f"Connection timeout while verifying character. Please contact {staff_mentions} for assistance."
    except requests.exceptions.RequestException as e:
        staff_mentions = "**Karumenta** or **Hokkies**"
        if guild:
            staff_mentions = await get_staff_mentions(guild)
        return False, f"Network error while verifying character. Please contact {staff_mentions} for assistance."
    except Exception as e:
        staff_mentions = "**Karumenta** or **Hokkies**"
        if guild:
            staff_mentions = await get_staff_mentions(guild)
        return False, f"Unexpected error while verifying character. Please contact {staff_mentions} for assistance."

class ApplicationView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)  # Persistent view
    
    @discord.ui.button(label='Apply', style=discord.ButtonStyle.green, emoji='📝')
    async def apply_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        user = interaction.user
        
        # Check if user already has Trial, Raider, Officer, or Guild Leader role
        guild = interaction.guild
        member = guild.get_member(user.id)
        if member:
            trial_role = discord.utils.get(guild.roles, name="Trial")
            raider_role = discord.utils.get(guild.roles, name="Raider")
            officer_role = discord.utils.get(guild.roles, name="Officer")
            guild_leader_role = discord.utils.get(guild.roles, name="Guild Leader")
            
            # Get staff mentions
            staff_mentions = await get_staff_mentions(guild)
            
            if trial_role and trial_role in member.roles:
                await interaction.response.send_message(f"❌ You already have the **Trial** role and cannot apply again. If you need assistance, please contact {staff_mentions}.", ephemeral=True)
                return
            
            if raider_role and raider_role in member.roles:
                await interaction.response.send_message(f"❌ You already have the **Raider** role and cannot apply again. If you need assistance, please contact {staff_mentions}.", ephemeral=True)
                return
            
            if officer_role and officer_role in member.roles:
                await interaction.response.send_message(f"❌ You already have the **Officer** role and cannot apply again. If you need assistance, please contact {staff_mentions}.", ephemeral=True)
                return
            
            if guild_leader_role and guild_leader_role in member.roles:
                await interaction.response.send_message(f"❌ You already have the **Guild Leader** role and cannot apply again. If you need assistance, please contact {staff_mentions}.", ephemeral=True)
                return
        
        # Check if user already has an active application
        if user.id in active_applications:
            await interaction.response.send_message("❌ You already have an active application in progress. Please complete it first or wait for it to expire.", ephemeral=True)
            return
        
        try:
            # Initialize application data
            current_time = asyncio.get_event_loop().time()
            active_applications[user.id] = {
                'question_index': -1,  # Start at -1 to indicate path selection phase
                'answers': [],
                'guild_id': interaction.guild.id,
                'start_time': current_time,
                'last_activity': current_time,
                'path': None,
                'questions': None,
                'warning_sent': False
            }
            
            # Send path selection
            embed = discord.Embed(
                title="🎉 Welcome to <Hopium>!",
                description=APPLICATION_CONFIG["intro"]["question"],
                color=0x00ff00
            )
            
            # Add option descriptions
            for option in APPLICATION_CONFIG["intro"]["options"]:
                embed.add_field(
                    name=option["label"],
                    value=option["description"],
                    inline=False
                )
            
            embed.set_footer(text="Please click one of the buttons below to select your application type.")
            
            view = ApplicationPathView(user.id, interaction.guild.id)
            await user.send(embed=embed, view=view)
            
            # Respond to the interaction
            await interaction.response.send_message("✅ Check your DMs! I've started your application process.", ephemeral=True)
            
        except discord.Forbidden:
            # User has DMs disabled
            await interaction.response.send_message("❌ I couldn't send you a DM. Please enable DMs from server members and try again.", ephemeral=True)
        except Exception as e:
            await interaction.response.send_message("❌ An error occurred. Please try again later or reach someone from the Staff.", ephemeral=True)
            print(f"Error sending DM: {e}")

class ApplicationPathView(discord.ui.View):
    def __init__(self, user_id, guild_id):
        super().__init__(timeout=PATH_SELECTION_TIMEOUT)  # 20 minute timeout
        self.user_id = user_id
        self.guild_id = guild_id
        
        # Add buttons for each path option
        for option in APPLICATION_CONFIG["intro"]["options"]:
            button = discord.ui.Button(
                label=option["label"],
                style=discord.ButtonStyle.primary,
                custom_id=f"path_{option['id']}"
            )
            button.callback = self.create_callback(option['id'])
            self.add_item(button)
    
    def create_callback(self, path_id):
        async def callback(interaction):
            if interaction.user.id != self.user_id:
                await interaction.response.send_message("❌ This is not your application.", ephemeral=True)
                return
            
            # Handle 'none' option - end application
            if path_id == 'none':
                if self.user_id in active_applications:
                    del active_applications[self.user_id]
                
                embed = discord.Embed(
                    title="❌ Application Cancelled",
                    description="Your application has been cancelled. Feel free to apply again in the future if your raiding interests change.",
                    color=0x888888
                )
                await interaction.response.edit_message(embed=embed, view=None)
                return
            
            # Update application data with chosen path
            if self.user_id in active_applications:
                current_time = asyncio.get_event_loop().time()
                active_applications[self.user_id]['path'] = path_id
                active_applications[self.user_id]['questions'] = APPLICATION_CONFIG["paths"][path_id]
                active_applications[self.user_id]['question_index'] = 0
                active_applications[self.user_id]['last_activity'] = current_time
                
                # Send first question of chosen path
                questions = APPLICATION_CONFIG["paths"][path_id]
                embed = discord.Embed(
                    title="📝 Let's Begin!",
                    description=f"You've chosen the **{next(opt['label'] for opt in APPLICATION_CONFIG['intro']['options'] if opt['id'] == path_id)}** application.",
                    color=0x00ff00
                )
                embed.add_field(
                    name=f"Question 1/{len(questions)}",
                    value=questions[0],
                    inline=False
                )
                embed.set_footer(text="Please respond with your answer. Type 'cancel' to cancel the application.")
                
                await interaction.response.edit_message(embed=embed, view=None)
            else:
                await interaction.response.send_message("❌ Application session expired. Please start a new application.", ephemeral=True)
        
        return callback
    
    async def on_timeout(self):
        # Clean up expired application
        if self.user_id in active_applications:
            del active_applications[self.user_id]

class ReviewView(discord.ui.View):
    def __init__(self, user_id, character_name, application_channel, review_channel):
        super().__init__(timeout=None)
        self.user_id = user_id
        self.character_name = character_name
        self.application_channel = application_channel
        self.review_channel = review_channel
    
    @discord.ui.button(label='Accept', style=discord.ButtonStyle.green, emoji='✅')
    async def accept_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        guild = interaction.guild
        member = guild.get_member(self.user_id)
        
        if not member:
            await interaction.response.send_message("❌ User not found in server.", ephemeral=True)
            return
        
        try:
            # Get or create "Trial" role
            trial_role = discord.utils.get(guild.roles, name="Trial")
            if not trial_role:
                trial_role = await guild.create_role(name="Trial")
                print("Created 'Trial' role!")
            
            # Remove "Social" role if user has it
            social_role = discord.utils.get(guild.roles, name="Social")
            if social_role and social_role in member.roles:
                await member.remove_roles(social_role)
                print(f"Removed 'Social' role from {member.display_name}")
            
            # Give user the Trial role
            await member.add_roles(trial_role)
            
            # Get or create "Trials" category with proper permissions
            trials_category = discord.utils.get(guild.categories, name="Trials")
            if not trials_category:
                # Set up permissions for Trials category (Officer, Bot, Guild Leader only)
                overwrites = {
                    guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False),
                }
                
                # Add permissions for specific roles
                officer_role = discord.utils.get(guild.roles, name="Officer")
                guild_leader_role = discord.utils.get(guild.roles, name="Guild Leader")
                bot_member = guild.me  # The bot itself
                
                if officer_role:
                    overwrites[officer_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
                if guild_leader_role:
                    overwrites[guild_leader_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
                if bot_member:
                    overwrites[bot_member] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
                
                trials_category = await guild.create_category("Trials", overwrites=overwrites)
                print("Created 'Trials' category with Officer/Guild Leader/Bot permissions!")
            
            # Rename application channel and move to Trials category
            if self.application_channel:
                new_channel_name = f"trial-{self.character_name.lower().replace(' ', '-')}"
                # When moving to Trials, keep user access for the trial channel
                trial_overwrites = {
                    guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False),
                    member: discord.PermissionOverwrite(read_messages=True, send_messages=True, view_channel=True)
                }
                await self.application_channel.edit(name=new_channel_name, category=trials_category, overwrites=trial_overwrites)
            
            # Send acceptance message to application channel
            if self.application_channel:
                accept_embed = discord.Embed(
                    title="🎉 Application Accepted!",
                    description=f"Congratulations {member.mention}! Your application has been accepted and you've been given the **Trial** role.",
                    color=0x00ff00
                )
                accept_embed.add_field(
                    name="📋 General Information",
                    value="Just some general info we work on a no sign up based roster, post in ⁠⛔absence if you're going to miss a raid, so i won't roster you for that week, and i try to post the roster around wednesday in ⁠📒raid-assigments and the assignments will be updated before the raid.",
                    inline=False
                )
                #accept_embed.add_field(
                #    name="🎯 TMB Setup Required",
                #    value="Please create a character on https://thatsmybis.com/ and add him to the guild from the home page. Once you do it notify us, Thanks!",
                #    inline=False
                #)
                accept_embed.add_field(
                    name="⚙️ Addons Required",
                    value="Please make sure you install RCLC lootcouncil before heading into your first raid with us, we use this addon to distribute loot in our raids 🙂",
                    inline=False
                )
                await self.application_channel.send(embed=accept_embed)
            
            # Send confirmation message to interaction BEFORE deleting the review channel
            await interaction.response.send_message(f"✅ Application accepted! {member.mention} has been given the Trial role and the trial channel has been moved to the Trials category.", ephemeral=False)
            
            # Get or create "review-all" channel under Trials category
            review_all_channel = discord.utils.get(guild.channels, name="review-all")
            if not review_all_channel:
                # Create "review-all" channel with officer/guild leader permissions under Trials category
                overwrites = {
                    guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False),
                }
                
                # Add permissions for specific roles
                officer_role = discord.utils.get(guild.roles, name="Officer")
                guild_leader_role = discord.utils.get(guild.roles, name="Guild Leader")
                bot_member = guild.me  # The bot itself
                
                if officer_role:
                    overwrites[officer_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
                if guild_leader_role:
                    overwrites[guild_leader_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
                if bot_member:
                    overwrites[bot_member] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
                
                review_all_channel = await guild.create_text_channel("review-all", category=trials_category, overwrites=overwrites)
                print("Created 'review-all' channel with Officer/Guild Leader/Bot permissions in Trials category!")
            else:
                # If review-all channel exists but is not in Trials category, move it there
                if review_all_channel.category != trials_category:
                    await review_all_channel.edit(category=trials_category)
                    print("Moved existing 'review-all' channel to Trials category!")
            
            # Send confirmation message to review-all channel
            if review_all_channel:
                confirmation_embed = discord.Embed(
                    title="✅ Application Accepted",
                    description=f"**Staff Member:** {interaction.user.mention}\n**Applicant:** {member.mention} ({member.display_name})\n**Character Name:** {self.character_name}",
                    color=0x00ff00
                )
                confirmation_embed.add_field(
                    name="📁 Trial Channel",
                    value=f"Trial channel: {self.application_channel.mention if self.application_channel else 'N/A'}",
                    inline=False
                )
                confirmation_embed.add_field(
                    name="⚡ Actions Taken",
                    value="• Trial role assigned\n• Channel moved to Trials category\n• Review channel deleted",
                    inline=False
                )
                await review_all_channel.send(embed=confirmation_embed)
            
            # Delete the review channel LAST to avoid interaction errors
            if self.review_channel:
                try:
                    await self.review_channel.delete()
                    print(f"Deleted review channel: {self.review_channel.name}")
                except Exception as delete_error:
                    print(f"Error deleting review channel: {delete_error}")
            
        except Exception as e:
            try:
                if interaction.response.is_done():
                    await interaction.followup.send(f"❌ Error processing acceptance: {e}", ephemeral=True)
                else:
                    await interaction.response.send_message(f"❌ Error processing acceptance: {e}", ephemeral=True)
            except Exception:
                # If we can't send the error message, just log it
                print(f"Error accepting application: {e}")
    
    @discord.ui.button(label='Decline', style=discord.ButtonStyle.red, emoji='❌')
    async def decline_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        try:
            # Send decline message to application channel before deletion
            if self.application_channel:
                decline_embed = discord.Embed(
                    title="❌ Application Declined",
                    description=f"Unfortunately {self.character_name}, your application has been declined. You may reapply in the future.",
                    color=0xff0000
                )
                await self.application_channel.send(embed=decline_embed)
            
            await interaction.response.send_message("❌ Application declined. Review channel will be deleted.", ephemeral=False)
            
            # Delete the review channel
            if self.review_channel:
                await self.review_channel.delete()
            
        except Exception as e:
            await interaction.response.send_message(f"❌ Error processing decline: {e}", ephemeral=True)
            print(f"Error declining application: {e}")

@bot.event
async def on_message(message):
    # Ignore bot messages
    if message.author.bot:
        return
    
    # Check if this is a DM and user has an active application
    if isinstance(message.channel, discord.DMChannel) and message.author.id in active_applications:
        await handle_application_response(message)
        return
    
    # Process commands
    await bot.process_commands(message)

async def handle_application_response(message):
    user_id = message.author.id
    app_data = active_applications[user_id]
    
    # Check for cancel command
    if message.content.lower() == 'cancel':
        del active_applications[user_id]
        embed = discord.Embed(
            title="❌ Application Cancelled",
            description="Your application has been cancelled. You can start a new one anytime by clicking the Apply button again.",
            color=0xff0000
        )
        await message.channel.send(embed=embed)
        return

    # Handle 'skip' for optional question groups
    path = app_data.get('path', '')
    optional_indices = OPTIONAL_QUESTIONS.get(path, set())
    current_index = app_data['question_index']
    questions = app_data.get('questions', [])

    if message.content.lower() == 'skip' and current_index in optional_indices:
        # Fill N/A for all remaining optional questions in this group and advance
        while app_data['question_index'] in optional_indices:
            app_data['answers'].append("N/A")
            app_data['question_index'] += 1
        app_data['last_activity'] = asyncio.get_event_loop().time()
        app_data['warning_sent'] = False
        if app_data['question_index'] < len(questions):
            embed = discord.Embed(title="📝 Next Question", color=0x00ff00)
            embed.add_field(
                name=f"Question {app_data['question_index'] + 1}/{len(questions)}",
                value=questions[app_data['question_index']],
                inline=False
            )
            embed.set_footer(text="Please respond with your answer. Type 'cancel' to cancel the application.")
            await message.channel.send(embed=embed)
        else:
            await complete_application(message.author, app_data)
        return
    
    # Check if user is still in path selection phase
    if app_data['question_index'] == -1:
        await message.channel.send("🤔 Please use the buttons above to select your application type.")
        return
    
    # Check if path is selected
    if not app_data.get('path') or not app_data.get('questions'):
        await message.channel.send("❌ Something went wrong with your application. Please start over.")
        del active_applications[user_id]
        return
    
    # Get the current questions for this path
    questions = app_data['questions']
    
    # Special validation for character name (first question) - COMMENTED OUT
    # if app_data['question_index'] == 0:
    #     character_name = message.content.strip()
    #     
    #     # Validate character name
    #     guild = bot.get_guild(app_data['guild_id'])
    #     is_valid, error_message = await validate_character_name(character_name, guild)
    #     
    #     if not is_valid:
    #         # Send error message and ask for character name again
    #         error_embed = discord.Embed(
    #             title="❌ Character Validation Failed",
    #             description=error_message,
    #             color=0xff0000
    #         )
    #         error_embed.add_field(
    #             name=f"Question 1/{len(questions)}",
    #             value=questions[0],
    #             inline=False
    #         )
    #         error_embed.set_footer(text="Please provide the correct character name or type 'cancel' to cancel the application.")
    #         await message.channel.send(embed=error_embed)
    #         return  # Don't advance to next question, ask again

    # Save the answer
    app_data['answers'].append(message.content)
    app_data['question_index'] += 1
    
    # Update activity tracking and reset warning state
    app_data['last_activity'] = asyncio.get_event_loop().time()
    app_data['warning_sent'] = False  # Reset warning for next question
    
    # Check if we have more questions
    if app_data['question_index'] < len(questions):
        # Send next question
        next_index = app_data['question_index']
        is_optional = next_index in optional_indices
        embed = discord.Embed(
            title="📝 Next Question",
            color=0x00ff00
        )
        embed.add_field(
            name=f"Question {next_index + 1}/{len(questions)}",
            value=questions[next_index],
            inline=False
        )
        if is_optional:
            embed.set_footer(text="This section is optional — type 'skip' to skip the alt character info. Type 'cancel' to cancel the application.")
        else:
            embed.set_footer(text="Please respond with your answer. Type 'cancel' to cancel the application.")

        await message.channel.send(embed=embed)
    else:
        # Application completed
        await complete_application(message.author, app_data)

async def send_character_reviews(review_channel, app_data):
    """Send character detail reviews for all character names mentioned in the application"""
    if not review_channel or not app_data:
        return
    
    # Get the questions and answers
    questions = app_data.get('questions', [])
    answers = app_data.get('answers', [])
    
    if len(questions) != len(answers):
        return
    
    # Find all character name questions and their corresponding answers
    character_data = []
    
    for i, (question, answer) in enumerate(zip(questions, answers)):
        question_lower = question.lower()
        if 'character name' in question_lower:
            # Extract the type of character from the question
            character_type = "Character"
            if "[speedrun]" in question.lower():
                if "main" in question.lower():
                    character_type = "⚡ Speedrun Main"
                else:
                    character_type = "⚡ Speedrun Alt"
            elif "[chill run]" in question.lower():
                character_type = "😎 Chill Run"
            
            character_data.append({
                'name': answer.strip(),
                'type': character_type,
                'question_index': i + 1
            })
    
    if not character_data:
        # Fallback: use first answer as character name if no specific character questions found
        if answers:
            character_data.append({
                'name': answers[0].strip(),
                'type': "Character",
                'question_index': 1
            })
    
    # Send character reviews
    for char_data in character_data:
        character_name = char_data['name']
        character_type = char_data['type']
        
        if not character_name or character_name.lower() in ['', 'unknown', 'n/a', 'none']:
            continue
        
        # Create character review embed
        embed = discord.Embed(
            title=f"🔍 Character Review - {character_name}",
            description=f"**Type:** {character_type}",
            color=0x0099ff
        )
        
        # Add lookup links
        embed.add_field(
            name="🔗 Warcraft Logs",
            value=f"[View WCL Profile](https://fresh.warcraftlogs.com/character/eu/spineshatter/{character_name.replace(' ', '%20')})",
            inline=True
        )
        embed.add_field(
            name="⚔️ Classic WoW Armory",
            value=f"[View Armory Profile](https://classicwowarmory.com/character/eu/spineshatter/{character_name.replace(' ', '%20')})",
            inline=True
        )
        
        # Check if character exists and add validation info
        character_exists, error_msg = await validate_character_exists(character_name)
        if not character_exists and error_msg:
            embed.add_field(
                name="⚠️ Validation Status",
                value=f"Character validation: {error_msg}",
                inline=False
            )
        else:
            embed.add_field(
                name="✅ Validation Status",
                value="Character found on armory",
                inline=False
            )
        
        embed.set_footer(text=f"Question {char_data['question_index']}: {questions[char_data['question_index']-1] if char_data['question_index']-1 < len(questions) else 'Unknown'}")
        
        await review_channel.send(embed=embed)
        
        # Small delay between character reviews to avoid rate limits
        await asyncio.sleep(0.5)

async def complete_application(user, app_data):
    # Remove from active applications
    del active_applications[user.id]
    
    # Get the guild
    guild = bot.get_guild(app_data['guild_id'])
    if not guild:
        return
    
    # Get the user's member object in the guild
    member = guild.get_member(user.id)
    if not member:
        return
    
    # Get the character name from the first answer
    character_name = app_data['answers'][0] if app_data['answers'] else "Unknown"
    
    # Store the old nickname for comparison
    old_nick = member.display_name
    
    # Rename the user to their character name
    try:
        await member.edit(nick=character_name)
        print(f"Renamed {user.display_name} to {character_name}")
    except discord.Forbidden:
        print(f"Cannot rename {user.display_name} - insufficient permissions")
    except Exception as e:
        print(f"Error renaming user: {e}")
    
    # Check if "Applications" category exists
    applications_category = discord.utils.get(guild.categories, name="Applications")
    
    if not applications_category:
        # Set up permissions for Applications category (Officer, Bot, Guild Leader only)
        overwrites = {
            guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False),
        }
        
        # Add permissions for specific roles
        officer_role = discord.utils.get(guild.roles, name="Officer")
        guild_leader_role = discord.utils.get(guild.roles, name="Guild Leader")
        bot_member = guild.me  # The bot itself
        
        if officer_role:
            overwrites[officer_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
        if guild_leader_role:
            overwrites[guild_leader_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
        if bot_member:
            overwrites[bot_member] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
        
        # Create the "Applications" category with permissions
        applications_category = await guild.create_category("Applications", overwrites=overwrites)
        print("Created 'Applications' category with Officer/Guild Leader/Bot permissions!")
    
    # Create application channel with user access (inherits category permissions + user access)
    application_channel_name = f"application-{character_name.lower().replace(' ', '-')}"
    try:
        # Set permissions for the application channel (inherits from category + user access)
        overwrites = {
            guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False),
            member: discord.PermissionOverwrite(read_messages=True, send_messages=True, view_channel=True)
        }
        
        application_channel = await guild.create_text_channel(
            application_channel_name,
            category=applications_category,
            overwrites=overwrites
        )
        print(f"Created application channel: {application_channel_name}")
    except Exception as e:
        print(f"Error creating application channel: {e}")
        application_channel = None
    
    # Create review channel (inherits category permissions only)
    review_channel_name = f"review-{character_name.lower().replace(' ', '-')}"
    try:
        # Set explicit permissions for review channel (deny @everyone, inherit officer/guild leader from category)
        overwrites = {
            guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False)
        }
        
        review_channel = await guild.create_text_channel(
            review_channel_name,
            category=applications_category,
            overwrites=overwrites
        )
        print(f"Created review channel: {review_channel_name}")
    except Exception as e:
        print(f"Error creating review channel: {e}")
        review_channel = None
    
    # Send completion message to user
    embed = discord.Embed(
        title="🎉 Application Completed!",
        description=f"Thank you for completing your application, {character_name}! Our staff will review it and get back to you soon.",
        color=0x00ff00
    )
    if application_channel:
        embed.add_field(
            name="Your Application Channel",
            value=f"You can check the status of your application in {application_channel.mention}",
            inline=False
        )
    await user.send(embed=embed)
    
    # Post the application in the application channel
    if application_channel:
        embed = discord.Embed(
            title=f"📋 Application for {character_name}",
            description=f"Application submitted by {user.mention}",
            color=0x0099ff
        )
        
        # Get the questions from the chosen path
        questions = app_data.get('questions', APPLICATION_QUESTIONS)  # Fallback to legacy questions
        path_name = app_data.get('path', 'unknown')
        
        # Add path information
        if path_name != 'unknown':
            path_label = next((opt['label'] for opt in APPLICATION_CONFIG['intro']['options'] if opt['id'] == path_name), path_name)
            embed.description = f"Application submitted by {user.mention}\n**Application for {path_label}**"
        
        for i, (question, answer) in enumerate(zip(questions, app_data['answers'])):
            embed.add_field(
                name=f"Q{i+1}: {question}",
                value=answer[:1024] if len(answer) <= 1024 else answer[:1021] + "...",
                inline=False
            )
        
        await application_channel.send(embed=embed)
        
        # Send nick change notification if nickname was different
        if old_nick != character_name:
            nick_embed = discord.Embed(
                title="📝 Nickname Updated",
                description=f"Your server nickname has been changed from **{old_nick}** to **{character_name}**",
                color=0x00ff00
            )
            await application_channel.send(embed=nick_embed)
    
    # Send review message with Accept/Decline buttons to review channel for staff
    if review_channel:
        embed = discord.Embed(
            title=f"📋 Review Application - {character_name}",
            description=f"**Applicant:** {user.mention} ({user.display_name})\n**Character Name:** {character_name}",
            color=0xffa500
        )
        embed.add_field(
            name="📁 Application Details",
            value=f"Full application can be viewed in {application_channel.mention if application_channel else 'application channel'}",
            inline=False
        )
        embed.add_field(
            name="⚡ Actions",
            value="Click **Accept** to give Trial role and move to Trials category\nClick **Decline** to reject and delete this review channel",
            inline=False
        )
        
        view = ReviewView(user.id, character_name, application_channel, review_channel)
        await review_channel.send(embed=embed, view=view)
        
        # Send character detail reviews for all characters mentioned in the application
        await send_character_reviews(review_channel, app_data)

@bot.event
@bot.event
async def on_ready():
    print(f'Logged in as {bot.user.name} - {bot.user.id}')
    logger.info(f'Bot logged in as {bot.user.name} - {bot.user.id}')
    print(f'Bot is in {len(bot.guilds)} guilds')
    logger.info(f'Bot is in {len(bot.guilds)} guilds')
    print('------')
    
    # Start the periodic tasks
    if not periodic_task.is_running():
        periodic_task.start()
        logger.info('Periodic data update task started')
    
    if not timeout_checker_task.is_running():
        timeout_checker_task.start()
        logger.info('Timeout checker task started')
    
    # Auto-run setupHopium for all guilds
    print(f"🔧 Running setupHopium for {len(bot.guilds)} guilds...")
    for guild in bot.guilds:
        try:
            # Create a mock context object to pass to setupHopium
            class MockContext:
                def __init__(self, guild):
                    self.guild = guild
                    self.message = None
                
                async def send(self, content):
                    # Just log instead of sending messages
                    print(f"  {guild.name}: {content}")
                    return self  # Return self to act as a message object
                
                async def delete(self):
                    # Mock delete method
                    pass
            
            mock_ctx = MockContext(guild)
            print(f"🔧 Setting up guild: {guild.name}")
            await setupHopium(mock_ctx)
            await asyncio.sleep(1)  # Small delay between guilds to avoid rate limits
        except Exception as e:
            print(f"❌ Failed to setup guild {guild.name}: {e}")
            logger.error(f"Failed to setup guild {guild.name}: {e}", exc_info=True)
    
    print("✅ setupHopium completed for all guilds!")
    logger.info("setupHopium completed for all guilds")

@bot.event
async def on_error(event, *args, **kwargs):
    """Handle general bot errors"""
    print(f'Error in event {event}: {args}')

@bot.event  
async def on_command_error(ctx, error):
    """Handle command errors"""
    if isinstance(error, commands.CommandNotFound):
        return  # Ignore unknown commands
    elif isinstance(error, commands.MissingPermissions):
        await ctx.send("❌ You don't have permission to use this command.")
    else:
        await ctx.send(f"❌ An error occurred: {str(error)}")
        print(f'Command error: {error}')

class BotManagementView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=None)  # Persistent view
    
    @discord.ui.button(label='Get Attendance', style=discord.ButtonStyle.primary, emoji='📊')
    async def get_attendance_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        # Check permissions - only allow Officers, Guild Leaders
        authorized_roles = ["Officer", "Guild Leader"]
        user_roles = [role.name for role in interaction.user.roles]
        
        if not any(role in authorized_roles for role in user_roles):
            await interaction.response.send_message("❌ You don't have permission to use this feature. Required roles: Officer or Guild Leader.", ephemeral=True)
            return
        
        try:
            # Get guild-specific data
            guild_id = interaction.guild.id
            paths = get_guild_file_paths(guild_id)
            
            # Initialize guild data if needed
            initialize_guild_data_files(guild_id)
            
            # Generate Excel for this guild
            await interaction.response.send_message("📊 Generating attendance report for this server...", ephemeral=True)
            
            # Create guild-specific Excel (Attendance type)
            workbook = createExcel(guild_id, "Attendance")
            
            if workbook:
                # Save and send the file
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"attendance_sheet_{interaction.guild.name}_{timestamp}.xlsx"
                file_path = os.path.join(paths['sheet_dir'], filename)
                
                workbook.save(file_path)
                
                # Edit the original message to send the file
                await interaction.edit_original_response(
                    content=f"📊 **Attendance Report for {interaction.guild.name}**\nGenerated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\nThis report contains attendance data for all guild members.",
                    attachments=[discord.File(file_path, filename=filename)]
                )
                
                # Clean up file
                os.remove(file_path)
            else:
                await interaction.edit_original_response(content="❌ Failed to generate Excel file. Please check the data files.")
            
        except Exception as e:
            await interaction.edit_original_response(content=f"❌ Error generating Excel report: {str(e)[:100]}...")
            logger.error(f"Error in attendance button: {e}", exc_info=True)
    
    @discord.ui.button(label='Get Role Items', style=discord.ButtonStyle.primary, emoji='⚔️')
    async def get_class_items_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        # Check permissions - only allow Officers, Guild Leaders
        authorized_roles = ["Officer", "Guild Leader"]
        user_roles = [role.name for role in interaction.user.roles]
        
        if not any(role in authorized_roles for role in user_roles):
            await interaction.response.send_message("❌ You don't have permission to use this feature. Required roles: Officer or Guild Leader.", ephemeral=True)
            return
        
        try:
            # Get guild-specific data
            guild_id = interaction.guild.id
            paths = get_guild_file_paths(guild_id)
            
            # Initialize guild data if needed
            initialize_guild_data_files(guild_id)
            
            # Generate Excel for this guild
            await interaction.response.send_message("⚔️ Generating class items report for this server...", ephemeral=True)
            
            # Create guild-specific Excel (Class Items type)
            workbook = createExcel(guild_id, "Class Items")
            
            if workbook:
                # Save and send the file
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"class_items_sheet_{interaction.guild.name}_{timestamp}.xlsx"
                file_path = os.path.join(paths['sheet_dir'], filename)
                
                workbook.save(file_path)
                
                # Edit the original message to send the file
                await interaction.edit_original_response(
                    content=f"⚔️ **Class Items Report for {interaction.guild.name}**\nGenerated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\nThis report contains class-specific item data and recommendations.",
                    attachments=[discord.File(file_path, filename=filename)]
                )
                
                # Clean up file
                os.remove(file_path)
            else:
                await interaction.edit_original_response(content="❌ Failed to generate Excel file. Please check the data files.")
            
        except Exception as e:
            await interaction.edit_original_response(content=f"❌ Error generating class items report: {str(e)[:100]}...")
            logger.error(f"Error in class items button: {e}", exc_info=True)
    
    @discord.ui.button(label='Get Loot', style=discord.ButtonStyle.primary, emoji='💎')
    async def get_loot_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        # Check permissions - only allow Officers, Guild Leaders
        authorized_roles = ["Officer", "Guild Leader"]
        user_roles = [role.name for role in interaction.user.roles]
        
        if not any(role in authorized_roles for role in user_roles):
            await interaction.response.send_message("❌ You don't have permission to use this feature. Required roles: Officer or Guild Leader.", ephemeral=True)
            return
        
        try:
            # Get guild-specific data
            guild_id = interaction.guild.id
            paths = get_guild_file_paths(guild_id)
            
            # Initialize guild data if needed
            initialize_guild_data_files(guild_id)
            
            # Generate Excel for this guild
            await interaction.response.send_message("💎 Generating loot distribution report for this server...", ephemeral=True)
            
            # Create guild-specific Excel (Loot type)
            workbook = createExcel(guild_id, "Loot")
            
            if workbook:
                # Save and send the file
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"loot_template_sheet_{interaction.guild.name}_{timestamp}.xlsx"
                file_path = os.path.join(paths['sheet_dir'], filename)
                
                workbook.save(file_path)
                
                # Edit the original message to send the file
                await interaction.edit_original_response(
                    content=f"💎 **Loot Distribution Report for {interaction.guild.name}**\nGenerated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\nThis report contains guild loot distribution data and analytics.",
                    attachments=[discord.File(file_path, filename=filename)]
                )
                
                # Clean up file
                os.remove(file_path)
            else:
                await interaction.edit_original_response(content="❌ Failed to generate Excel file. Please check the data files.")
            
        except Exception as e:
            await interaction.edit_original_response(content=f"❌ Error generating loot report: {str(e)[:100]}...")
            logger.error(f"Error in loot button: {e}", exc_info=True)
    
    @discord.ui.button(label='Get All', style=discord.ButtonStyle.success, emoji='📦')
    async def get_all_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        # Check permissions - only allow Officers, Guild Leaders
        authorized_roles = ["Officer", "Guild Leader"]
        user_roles = [role.name for role in interaction.user.roles]
        
        if not any(role in authorized_roles for role in user_roles):
            await interaction.response.send_message("❌ You don't have permission to use this feature. Required roles: Officer or Guild Leader.", ephemeral=True)
            return
        
        # This button does the same as Get Attendance for now (generates the main Excel file)
        try:
            # Get guild-specific data
            guild_id = interaction.guild.id
            paths = get_guild_file_paths(guild_id)
            
            # Initialize guild data if needed
            initialize_guild_data_files(guild_id)
            
            # Generate Excel for this guild
            await interaction.response.send_message("📦 Generating comprehensive guild data export for this server...", ephemeral=True)
            
            # Create guild-specific Excel (All types)
            workbook = createExcel(guild_id, "All")
            
            if workbook:
                # Save and send the file
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"complete_guild_export_{interaction.guild.name}_{timestamp}.xlsx"
                file_path = os.path.join(paths['sheet_dir'], filename)
                
                workbook.save(file_path)
                
                # Edit the original message to send the file
                await interaction.edit_original_response(
                    content=f"📦 **Complete Guild Data Export for {interaction.guild.name}**\nGenerated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\nThis file contains all available guild data including attendance, armory items, and performance statistics.",
                    attachments=[discord.File(file_path, filename=filename)]
                )
                
                # Clean up file
                os.remove(file_path)
            else:
                await interaction.edit_original_response(content="❌ Failed to generate Excel file. Please check the data files.")
            
        except Exception as e:
            await interaction.edit_original_response(content=f"❌ Error generating comprehensive export: {str(e)[:100]}...")
            logger.error(f"Error in get all button: {e}", exc_info=True)

@bot.command()
async def setupHopium(ctx):
    guild = ctx.guild
    setup_messages = []  # Track messages to delete later
    
    # Check if "Recruitment" category exists
    recruitment_category = discord.utils.get(guild.categories, name="Recruitment")
    
    if not recruitment_category:
        # Create the "Recruitment" category with restricted permissions
        overwrites = {
            guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False)
        }
        recruitment_category = await guild.create_category("Recruitment", overwrites=overwrites)
        msg = await ctx.send("Created 'Recruitment' category!")
        setup_messages.append(msg)
    
    # Check if "apply-here" channel exists in the category
    apply_channel = discord.utils.get(recruitment_category.channels, name="✍apply-here")
    
    if not apply_channel:
        # Create the "apply-here" channel in the Recruitment category with restricted permissions
        overwrites = {
            guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False)
        }
        apply_channel = await guild.create_text_channel("✍apply-here", category=recruitment_category, overwrites=overwrites)
        msg = await ctx.send("Created 'apply-here' channel in the Recruitment category!")
        setup_messages.append(msg)
    
    # Clear the apply-here channel to ensure clean state
    await apply_channel.purge()
    msg = await ctx.send("Cleared 'apply-here' channel!")
    setup_messages.append(msg)
    
    # Create the application message with button
    embed = discord.Embed(
        title="📋 Application for Hopium Guild",
        description=f"Click the button below to start your application process!\nIf anything goes wrong, please contact {await get_staff_mentions(guild)}.",
        color=0x00ff00
    )
    
    view = ApplicationView()
    await apply_channel.send(embed=embed, view=view)
    msg = await ctx.send("Application message sent with Apply button!")
    setup_messages.append(msg)
    
    # Check if "ADMIN" category exists
    admin_category = discord.utils.get(guild.categories, name="ADMIN")
    
    if not admin_category:
        # Create the "ADMIN" category with restricted permissions (Officers and Guild Leaders only)
        overwrites = {
            guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False)
        }
        
        # Add permissions for specific roles
        officer_role = discord.utils.get(guild.roles, name="Officer")
        guild_leader_role = discord.utils.get(guild.roles, name="Guild Leader")
        bot_member = guild.me  # The bot itself
        
        if officer_role:
            overwrites[officer_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
        if guild_leader_role:
            overwrites[guild_leader_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
        if bot_member:
            overwrites[bot_member] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
        
        admin_category = await guild.create_category("ADMIN", overwrites=overwrites)
        msg = await ctx.send("Created 'ADMIN' category!")
        setup_messages.append(msg)
    
    # Check if "HopiumBot" channel exists in the ADMIN category
    hopium_bot_channel = discord.utils.get(admin_category.channels, name="🤖hopiumbot")
    
    if not hopium_bot_channel:
        # Create the "HopiumBot" channel in the ADMIN category (inherits category permissions)
        overwrites = {
            guild.default_role: discord.PermissionOverwrite(read_messages=False, send_messages=False, view_channel=False)
        }
        
        # Add permissions for specific roles
        officer_role = discord.utils.get(guild.roles, name="Officer")
        guild_leader_role = discord.utils.get(guild.roles, name="Guild Leader")
        bot_member = guild.me
        
        if officer_role:
            overwrites[officer_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
        if guild_leader_role:
            overwrites[guild_leader_role] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
        if bot_member:
            overwrites[bot_member] = discord.PermissionOverwrite(read_messages=True, send_messages=True, manage_messages=True, view_channel=True)
        
        hopium_bot_channel = await guild.create_text_channel("🤖hopiumbot", category=admin_category, overwrites=overwrites)
        msg = await ctx.send("Created 'HopiumBot' channel in the ADMIN category!")
        setup_messages.append(msg)
    
    # Clear the HopiumBot channel to ensure clean state
    await hopium_bot_channel.purge()
    msg = await ctx.send("Cleared 'HopiumBot' channel!")
    setup_messages.append(msg)
    
    # Create the bot guide message with management buttons
    guide_embed = discord.Embed(
        title="🤖 HopiumBot Management Panel",
        description="Welcome to the HopiumBot management interface! Use the buttons below for quick access to guild data or the commands listed for advanced operations.",
        color=0x9932cc
    )
    
    guide_embed.add_field(
        name="📋 File Download Commands",
        value="• `!getfile armory` - Download guild armory data\n• `!getfile icons` - Download item icons data\n• `!getfile parses` - Download guild WCL parses data\n• `!getfile tmb` - Download TMB files (character, attendance, item notes)",
        inline=False
    )
    
    guide_embed.add_field(
        name="👤 Player Data Commands", 
        value="• `!get armory <playerName>` - Get specific player's armory data\n• `!get parses <playerName>` - Get specific player's WCL parses",
        inline=False
    )
    
    guide_embed.add_field(
        name="📤 Upload Commands",
        value="• `!uploadtmb` - Upload TMB files (character-json.json, hopium-attendance.csv, item-notes.csv)\n• `!uploadarmory` - Upload armory.json file (merges with existing data)",
        inline=False
    )
    
    guide_embed.add_field(
        name="⚙️ Management Commands",
        value="• `!setupHopium` - Run initial bot setup (creates channels and categories)\n• All commands are restricted to Officers and Guild Leaders only",
        inline=False
    )
    
    guide_embed.add_field(
        name="📊 Excel Generation",
        value="Use the buttons below to generate Excel reports:\n• **Get Attendance** - Guild attendance sheet\n• **Get Role Items** - Role sheets (Tank/DPS/Casters/Healers) per raid group\n• **Get Loot** - Loot sheet per raid group\n• **Get All** - All of the above in one file",
        inline=False
    )
    
    guide_embed.set_footer(text="Click the buttons below for Excel generation features • Officers & Guild Leaders only")
    
    management_view = BotManagementView()
    await hopium_bot_channel.send(embed=guide_embed, view=management_view)
    msg = await ctx.send("Bot management panel created with guide and buttons!")
    setup_messages.append(msg)
    
    msg = await ctx.send('✅ Setup completed! Messages will be deleted in 5 seconds...')
    setup_messages.append(msg)
    
    # Wait 5 seconds then delete all setup messages including the command message
    await asyncio.sleep(5)
    
    # Delete the original command message
    try:
        if getattr(ctx, 'message', None) is not None:
            await ctx.message.delete()
    except discord.NotFound:
        pass
    
    # Delete all setup status messages
    for message in setup_messages:
        try:
            await message.delete()
        except discord.NotFound:
            pass

@bot.command(name='getfile')
async def get_file_data(ctx, data_type: str = None):
    """
    Get guild data files as temporary messages
    Usage: !getfile armory | !getfile icons
    Restricted to Officers, Guild Leaders, and authorized roles
    """
    # Check permissions - only allow Officers, Guild Leaders
    authorized_roles = ["Officer", "Guild Leader"]
    user_roles = [role.name for role in ctx.author.roles]
    
    if not any(role in authorized_roles for role in user_roles):
        await ctx.send("❌ You don't have permission to use this command. Required roles: Officer or Guild Leader.", delete_after=10)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    if data_type not in ['armory', 'icons', 'tmb', 'parses']:
        embed = discord.Embed(
            title="📋 Available Data Types",
            description=f"Choose from available data file types for **{ctx.guild.name}**:",
            color=0xff9900
        )
        embed.add_field(
            name="📥 Download Commands",
            value="`!getfile armory` - Download guild armory data\n`!getfile icons` - Download item icons data\n`!getfile parses` - Download guild WCL parses data\n`!getfile tmb` - Download TMB data files (character, attendance, item notes)",
            inline=False
        )
        embed.add_field(
            name="📤 Upload Commands",
            value="`!uploadtmb` - Upload TMB files (character-json.json, hopium-attendance.csv, item-notes.csv)\n`!uploadarmory` - Upload armory.json file (merges with existing data)",
            inline=False
        )
        embed.add_field(
            name="🏰 Server Info",
            value=f"**Guild:** {ctx.guild.name}\n**All data operations are server-specific**",
            inline=False
        )
        embed.set_footer(text=f"Guild: {ctx.guild.name} • Data is server-specific")
        await ctx.send(embed=embed, delete_after=15)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    # Get guild-specific file paths
    guild_id = ctx.guild.id
    paths = get_guild_file_paths(guild_id)
    
    # Initialize guild data if needed
    initialize_guild_data_files(guild_id)
    
    try:
        # Handle TMB files differently (zip archive)
        if data_type == 'tmb':
            # Define TMB files using guild-specific paths
            tmb_files = [
                (paths['character_file'], 'character-json.json'),
                (paths['attendance_file'], 'hopium-attendance.csv'),
                (paths['item_file'], 'item-notes.csv')
            ]
            
            # Check which files exist
            existing_files = []
            missing_files = []
            total_size = 0
            
            for file_path, filename in tmb_files:
                if os.path.exists(file_path):
                    existing_files.append((file_path, filename))
                    total_size += os.path.getsize(file_path)
                else:
                    missing_files.append(filename)
            
            if not existing_files:
                await ctx.send("❌ No TMB files found. Ensure the TMB directory contains data files.", delete_after=15)
                return
            
            # Create zip file in guild-specific directory
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            zip_filename = f"tmb_data_{ctx.guild.name}_{timestamp}.zip"
            zip_path = os.path.join(paths['sheet_dir'], zip_filename)
            
            try:
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file_path, filename in existing_files:
                        zipf.write(file_path, filename)
                
                # Create embed for TMB data
                embed = discord.Embed(
                    title="📊 Guild TMB Data",
                    description=f"TMB data files archive containing {len(existing_files)} files",
                    color=0x0066cc,
                    timestamp=datetime.now()
                )
                
                # Add files summary
                file_list = []
                for _, filename in existing_files:
                    file_list.append(f"✅ {filename}")
                
                if missing_files:
                    for filename in missing_files:
                        file_list.append(f"❌ {filename} (missing)")
                
                embed.add_field(
                    name="📋 Files Included",
                    value="\n".join(file_list),
                    inline=False
                )
                
                # Archive info
                zip_size = os.path.getsize(zip_path)
                embed.add_field(
                    name="📁 Archive Info",
                    value=f"**Archive Size:** {zip_size:,} bytes\n**Total Files:** {len(existing_files)}\n**Created:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    inline=True
                )
                
                embed.add_field(
                    name="👤 Requested by",
                    value=ctx.author.mention,
                    inline=True
                )
                
                embed.set_footer(text="This message will be deleted in 60 seconds • Data is sensitive")
                
                # Send the zip file
                await ctx.send(
                    embed=embed,
                    file=discord.File(zip_path, filename=zip_filename),
                    delete_after=60
                )
                
                # Clean up temporary zip file
                os.remove(zip_path)
                
                # Log the action
                print(f"🔒 TMB data downloaded by {ctx.author} ({ctx.author.id}) in {ctx.guild.name}")
                
            except Exception as e:
                # Clean up zip file if it was created
                if os.path.exists(zip_path):
                    os.remove(zip_path)
                raise e
        
        else:
            # Handle single files (armory, icons, parses) using guild-specific paths
            if data_type == 'armory':
                file_path = paths['armory_file']
                file_type = "Armory"
                icon = "🛡️"
            elif data_type == 'icons':
                file_path = paths['item_icons_file']
                file_type = "Item Icons"
                icon = "🖼️"
            elif data_type == 'parses':
                file_path = paths['parses_file']
                file_type = "WCL Parses"
                icon = "📊"
            
            # Check if file exists
            if not os.path.exists(file_path):
                await ctx.send(f"❌ {file_type} file not found. Run the periodic task first to generate data.", delete_after=15)
                return
            
            # Load data
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if not data:
                await ctx.send(f"ℹ️ {file_type} file is empty. No data available.", delete_after=15)
                return
            
            # Create embed with data summary based on type
            if data_type == 'armory':
                embed = discord.Embed(
                    title=f"{icon} Guild {file_type} Data",
                    description=f"Character equipment data from {len(data)} players",
                    color=0x00ff00,
                    timestamp=datetime.now()
                )
                
                # Add character summary
                total_items = sum(len(items) for items in data.values())
                embed.add_field(
                    name="📊 Summary",
                    value=f"**Players:** {len(data)}\n**Total Items Tracked:** {total_items}",
                    inline=False
                )
            elif data_type == 'icons':
                embed = discord.Embed(
                    title=f"{icon} Guild {file_type} Data",
                    description=f"Item icon data for {len(data)} items",
                    color=0x9932cc,
                    timestamp=datetime.now()
                )
                
                # Add icons summary
                embed.add_field(
                    name="📊 Summary",
                    value=f"**Total Items:** {len(data)}",
                    inline=False
                )
            elif data_type == 'parses':
                embed = discord.Embed(
                    title=f"{icon} Guild {file_type} Data",
                    description=f"WCL performance data for {len(data)} players",
                    color=0xff6600,
                    timestamp=datetime.now()
                )
                
                # Add parses summary
                valid_players = sum(1 for player_data in data.values() if player_data.get("bestPerformanceAverage", 0) > 0)
                avg_best = sum(player_data.get("bestPerformanceAverage", 0) for player_data in data.values()) / len(data) if data else 0
                avg_median = sum(player_data.get("medianPerformanceAverage", 0) for player_data in data.values()) / len(data) if data else 0
                
                embed.add_field(
                    name="📊 Summary",
                    value=f"**Total Players:** {len(data)}\n**Players with Data:** {valid_players}\n**Avg Best Performance:** {avg_best:.1f}\n**Avg Median Performance:** {avg_median:.1f}",
                    inline=False
                )
            
            # File info (common for both types)
            file_size = os.path.getsize(file_path)
            file_modified = datetime.fromtimestamp(os.path.getmtime(file_path))
            embed.add_field(
                name="📁 File Info",
                value=f"**Size:** {file_size:,} bytes\n**Last Updated:** {file_modified.strftime('%Y-%m-%d %H:%M:%S')}",
                inline=True
            )
            
            embed.add_field(
                name="👤 Requested by",
                value=ctx.author.mention,
                inline=True
            )
            
            embed.set_footer(text="This message will be deleted in 60 seconds • Data is sensitive")
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{data_type}_data_{timestamp}.json"
            
            # Send the file as attachment with embed
            message = await ctx.send(
                embed=embed, 
                file=discord.File(file_path, filename=filename),
                delete_after=60
            )
            
            # Log the action
            print(f"🔒 {file_type} data downloaded by {ctx.author} ({ctx.author.id}) in {ctx.guild.name}")
        
        # Delete the command message
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
            
    except json.JSONDecodeError:
        if data_type == 'tmb':
            await ctx.send("❌ Error: One of the TMB files contains invalid JSON data.", delete_after=15)
        else:
            await ctx.send(f"❌ Error: {data_type.title()} file is corrupted or contains invalid JSON.", delete_after=15)
    except FileNotFoundError:
        if data_type == 'tmb':
            await ctx.send("❌ Error: TMB files not found.", delete_after=15)
        else:
            await ctx.send(f"❌ Error: {data_type.title()} file not found.", delete_after=15)
    except Exception as e:
        await ctx.send(f"❌ Error retrieving {data_type} data: {str(e)[:100]}...", delete_after=15)
        print(f"Error in get_file_data command: {e}")

@bot.command(name='get')
async def get_player_data(ctx, data_type: str = None, player_name: str = None):
    """
    Get player armory or parses data as a temporary message
    Usage: !get armory <playerName> | !get parses <playerName>
    Restricted to Officers, Guild Leaders, and authorized roles
    """
    # Check permissions - only allow Officers, Guild Leaders
    authorized_roles = ["Officer", "Guild Leader"]
    user_roles = [role.name for role in ctx.author.roles]
    
    if not any(role in authorized_roles for role in user_roles):
        await ctx.send("❌ You don't have permission to use this command. Required roles: Officer or Guild Leader.", delete_after=10)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    if data_type not in ['armory', 'parses'] or not player_name:
        embed = discord.Embed(
            title="📋 Player Data Lookup",
            description=f"Use `!get <type> <playerName>` to retrieve specific player's data from **{ctx.guild.name}**.",
            color=0xff9900
        )
        embed.add_field(
            name="Valid Commands",
            value="`!get armory <playerName>` - Get specific player's items\n`!get parses <playerName>` - Get specific player's WCL performance data\n`!getfile armory` - Download full guild armory data\n`!getfile parses` - Download full guild parses data\n`!uploadarmory` - Upload armory.json file (merges with existing)",
            inline=False
        )
        embed.add_field(
            name="Examples",
            value="`!get armory Karumenta` - Get Karumenta's items\n`!get parses Karumenta` - Get Karumenta's performance data",
            inline=False
        )
        embed.add_field(
            name="🏰 Server Info",
            value=f"**Guild:** {ctx.guild.name}\n**Data is specific to this server**",
            inline=False
        )
        embed.set_footer(text=f"Guild: {ctx.guild.name} • Data is server-specific")
        await ctx.send(embed=embed, delete_after=15)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    # Get guild-specific file paths
    guild_id = ctx.guild.id
    paths = get_guild_file_paths(guild_id)
    
    # Initialize guild data if needed
    initialize_guild_data_files(guild_id)
    
    try:
        # Determine which file to use based on data type
        if data_type == 'armory':
            file_path = paths['armory_file']
            file_type = "Armory"
            icon = "🛡️"
        elif data_type == 'parses':
            file_path = paths['parses_file']
            file_type = "WCL Parses"
            icon = "📊"
        
        # Check if file exists
        if not os.path.exists(file_path):
            await ctx.send(f"❌ {file_type} file not found. Run the periodic task first to generate data.", delete_after=15)
            return
        
        # Load data
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if not data:
            await ctx.send(f"ℹ️ {file_type} file is empty. No character data available.", delete_after=15)
            return
        
        # Find the player (case-insensitive search)
        player_found = None
        player_data = None
        
        for char_name, char_data in data.items():
            if char_name.lower() == player_name.lower():
                player_found = char_name
                player_data = char_data
                break
        
        if not player_found:
            # Search for partial matches
            partial_matches = []
            for char_name in data.keys():
                if player_name.lower() in char_name.lower():
                    partial_matches.append(char_name)
            
            if partial_matches:
                embed = discord.Embed(
                    title="❓ Player Not Found - Did you mean?",
                    description=f"Could not find exact match for '{player_name}'. Did you mean one of these?",
                    color=0xffa500
                )
                
                suggestion_list = []
                for match in partial_matches[:10]:  # Limit to 10 suggestions
                    if data_type == 'armory':
                        data_count = len(data[match]) if isinstance(data[match], list) else 0
                        suggestion_list.append(f"**{match}** ({data_count} items)")
                    elif data_type == 'parses':
                        best_avg = data[match].get("bestPerformanceAverage", 0) if isinstance(data[match], dict) else 0
                        suggestion_list.append(f"**{match}** (Best: {best_avg:.1f})")
                
                embed.add_field(
                    name="Suggestions",
                    value="\n".join(suggestion_list),
                    inline=False
                )
                embed.set_footer(text="Use the exact character name for best results")
            else:
                embed = discord.Embed(
                    title="❌ Player Not Found",
                    description=f"No player found with name '{player_name}' in the {file_type.lower()} data.",
                    color=0xff0000
                )
                
            await ctx.send(embed=embed, delete_after=30)
            try:
                await ctx.message.delete()
            except discord.NotFound:
                pass
            return
        
        # Player found, create embed based on data type
        if data_type == 'armory':
            # Handle armory data (list of items)
            player_items = player_data if isinstance(player_data, list) else []
            
            if not player_items:
                embed = discord.Embed(
                    title=f"📦 {player_found}'s Armory",
                    description="No items found for this player.",
                    color=0xffa500
                )
            else:
                embed = discord.Embed(
                    title=f"🛡️ {player_found}'s Armory",
                    description=f"Found **{len(player_items)}** items for {player_found}",
                    color=0x00ff00,
                    timestamp=datetime.now()
                )
                
                # Format all items as a single list (no chunking)
                item_list = []
                for item in player_items:
                    item_list.append(f"• {item}")
                
                # Join all items into one field value, Discord will handle truncation if needed
                items_text = "\n".join(item_list)
                
                # If the text is too long for a single field, we'll truncate with a note
                if len(items_text) > 1024:
                    # Find a good truncation point (at a line break)
                    truncate_at = items_text.rfind("\n", 0, 1000)
                    if truncate_at == -1:
                        truncate_at = 1000
                    items_text = items_text[:truncate_at] + f"\n... and {len(player_items) - items_text[:truncate_at].count('•')} more items"
                
                embed.add_field(
                    name="🎽 Equipment",
                    value=items_text,
                    inline=False
                )
        
        elif data_type == 'parses':
            # Handle parses data (dictionary with performance metrics)
            if not isinstance(player_data, dict) or not player_data:
                embed = discord.Embed(
                    title=f"📊 {player_found}'s WCL Parses",
                    description="No parse data found for this player.",
                    color=0xffa500
                )
            else:
                # Extract parse data
                best_avg = player_data.get("bestPerformanceAverage", 0)
                median_avg = player_data.get("medianPerformanceAverage", 0)
                
                # Determine color based on performance
                if best_avg >= 95:
                    color = 0xff6600  # Orange (Legendary)
                elif best_avg >= 75:
                    color = 0x9d4edd  # Purple (Epic)
                elif best_avg >= 50:
                    color = 0x0099ff  # Blue (Rare)
                elif best_avg >= 25:
                    color = 0x00ff00  # Green (Uncommon)
                else:
                    color = 0x808080  # Gray (Poor)
                
                embed = discord.Embed(
                    title=f"📊 {player_found}'s WCL Parses",
                    description=f"Performance data for {player_found}",
                    color=color,
                    timestamp=datetime.now()
                )
                
                emoji = "🏆"
                # Performance rating
                if best_avg >= 95:
                    emoji = "🧡"
                elif best_avg >= 75:
                    emoji = "💜"
                elif best_avg >= 50:
                    emoji = "💙 re"
                elif best_avg >= 25:
                    emoji = "💚"
                else:
                    emoji = "🤍"

                # Performance metrics
                embed.add_field(
                    name= emoji+" Performance Averages",
                    value=f"**Best Performance:** {best_avg:.1f}\n**Median Performance:** {median_avg:.1f}",
                    inline=False
                )
                
                # Add any additional parse data if available
                if "encounters" in player_data and isinstance(player_data["encounters"], dict):
                    encounter_list = []
                    for encounter, data in list(player_data["encounters"].items())[:5]:  # Show top 5 encounters
                        if isinstance(data, dict) and "bestPercent" in data:
                            encounter_list.append(f"• **{encounter}**: {data['bestPercent']:.1f}%")
                    
                    if encounter_list:
                        embed.add_field(
                            name="🎯 Top Encounters",
                            value="\n".join(encounter_list),
                            inline=False
                        )
        
        # Add metadata
        file_modified = datetime.fromtimestamp(os.path.getmtime(file_path))
        embed.add_field(
            name="📊 Info",
            value=f"**Last Updated:** {file_modified.strftime('%Y-%m-%d %H:%M:%S')}\n**Requested by:** {ctx.author.mention}",
            inline=False
        )
        
        embed.set_footer(text="This message will be deleted in 45 seconds")
        
        # Send the embed
        await ctx.send(embed=embed, delete_after=45)
        
        # Log the action
        print(f"🔍 Player {file_type.lower()} lookup: {player_found} by {ctx.author} ({ctx.author.id}) in {ctx.guild.name}")
        
        # Delete the command message
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
            
    except json.JSONDecodeError:
        await ctx.send(f"❌ Error: {file_type} file is corrupted or contains invalid JSON.", delete_after=15)
    except FileNotFoundError:
        await ctx.send(f"❌ Error: {file_type} file not found.", delete_after=15)
    except Exception as e:
        await ctx.send(f"❌ Error retrieving player {file_type.lower()} data: {str(e)[:100]}...", delete_after=15)
        print(f"Error in get_player_armory command: {e}")

@bot.command(name='uploadtmb')
async def upload_tmb_files(ctx):
    """
    Upload TMB files (character-json.json, hopium-attendance.csv, item-notes.csv)
    Usage: !uploadtmb (attach 1-3 files)
    Restricted to Officers, Guild Leaders, and authorized roles
    """
    # Check permissions - only allow Officers, Guild Leaders
    authorized_roles = ["Officer", "Guild Leader"]
    user_roles = [role.name for role in ctx.author.roles]
    
    if not any(role in authorized_roles for role in user_roles):
        await ctx.send("❌ You don't have permission to use this command. Required roles: Officer or Guild Leader.", delete_after=10)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    # Get guild-specific file paths
    guild_id = ctx.guild.id
    paths = get_guild_file_paths(guild_id)
    
    # Initialize guild data if needed
    initialize_guild_data_files(guild_id)
    
    # Check if files are attached
    if not ctx.message.attachments:
        embed = discord.Embed(
            title="📤 TMB File Upload",
            description=f"Upload TMB data files to update the **{ctx.guild.name}** guild database.",
            color=0xff9900
        )
        embed.add_field(
            name="📋 Supported Files",
            value="• `character-json.json` - Character data\n• `hopium-attendance.csv` - Attendance records\n• `item-notes.csv` - Item notes",
            inline=False
        )
        embed.add_field(
            name="📝 Instructions",
            value="1. Attach 1-3 files to your message\n2. Use the `!uploadtmb` command\n3. Files will be validated before overwriting",
            inline=False
        )
        embed.add_field(
            name="⚠️ Important",
            value="Only files with matching names will be updated. Invalid files will be rejected.",
            inline=False
        )
        embed.add_field(
            name="🏰 Server Info",
            value=f"**Guild:** {ctx.guild.name}\n**Files will be uploaded to this server's data only**",
            inline=False
        )
        embed.set_footer(text=f"Guild: {ctx.guild.name} • Data is server-specific")
        await ctx.send(embed=embed, delete_after=30)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    # Validate file count
    if len(ctx.message.attachments) > 3:
        await ctx.send("❌ Too many files attached. Maximum 3 files allowed (character-json.json, hopium-attendance.csv, item-notes.csv).", delete_after=15)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    try:
        # Define valid TMB files using guild-specific paths
        valid_files = {
            'character-json.json': (paths['character_file'], 'json'),
            'hopium-attendance.csv': (paths['attendance_file'], 'csv'),
            'item-notes.csv': (paths['item_file'], 'csv')
        }
        
        processed_files = []
        validation_errors = []
        uploaded_files = []
        
        # Process each attachment
        for attachment in ctx.message.attachments:
            filename = attachment.filename.lower()
            
            # Check if filename is valid
            if filename not in valid_files:
                validation_errors.append(f"❌ **{attachment.filename}** - Invalid filename. Expected: {', '.join(valid_files.keys())}")
                continue
            
            # Check file size (max 10MB)
            if attachment.size > 10 * 1024 * 1024:
                validation_errors.append(f"❌ **{attachment.filename}** - File too large (max 10MB)")
                continue
            
            target_path, file_type = valid_files[filename]
            
            try:
                # Download file content
                file_content = await attachment.read()
                
                # Validate file content based on type
                if file_type == 'json':
                    try:
                        # Validate JSON structure
                        json_data = json.loads(file_content.decode('utf-8'))
                        
                        # Additional validation for character-json.json
                        if filename == 'character-json.json':
                            if not isinstance(json_data, list):
                                validation_errors.append(f"❌ **{attachment.filename}** - Invalid format: Expected JSON array")
                                continue
                            
                            # Validate each character entry
                            for i, entry in enumerate(json_data):
                                if not isinstance(entry, dict):
                                    validation_errors.append(f"❌ **{attachment.filename}** - Invalid character entry at index {i}")
                                    break
                                if 'name' not in entry:
                                    validation_errors.append(f"❌ **{attachment.filename}** - Missing 'name' field in character entry at index {i}")
                                    break
                            else:
                                # All entries valid
                                processed_files.append((target_path, file_content, attachment.filename))
                        else:
                            # Generic JSON validation passed
                            processed_files.append((target_path, file_content, attachment.filename))
                            
                    except json.JSONDecodeError as e:
                        validation_errors.append(f"❌ **{attachment.filename}** - Invalid JSON format: {str(e)[:100]}")
                        continue
                    except UnicodeDecodeError:
                        validation_errors.append(f"❌ **{attachment.filename}** - Invalid encoding, expected UTF-8")
                        continue
                
                elif file_type == 'csv':
                    try:
                        # Validate CSV structure
                        csv_content = file_content.decode('utf-8')
                        csv_lines = csv_content.strip().split('\n')
                        
                        if not csv_lines or not csv_lines[0].strip():
                            validation_errors.append(f"❌ **{attachment.filename}** - Empty CSV file")
                            continue
                        
                        # Basic CSV validation - check if it can be parsed
                        import io
                        reader = csv.reader(io.StringIO(csv_content))
                        row_count = 0
                        for row in reader:
                            row_count += 1
                            if row_count > 1000:  # Prevent excessive processing
                                break
                        
                        if row_count == 0:
                            validation_errors.append(f"❌ **{attachment.filename}** - No data found in CSV")
                            continue
                        
                        processed_files.append((target_path, file_content, attachment.filename))
                        
                    except UnicodeDecodeError:
                        validation_errors.append(f"❌ **{attachment.filename}** - Invalid encoding, expected UTF-8")
                        continue
                    except Exception as e:
                        validation_errors.append(f"❌ **{attachment.filename}** - CSV validation error: {str(e)[:100]}")
                        continue
                
            except Exception as e:
                validation_errors.append(f"❌ **{attachment.filename}** - Download error: {str(e)[:100]}")
                continue
        
        # Check if any files were processed successfully
        if not processed_files and validation_errors:
            # All files failed validation
            embed = discord.Embed(
                title="❌ Upload Failed",
                description="All files failed validation. No files were updated.",
                color=0xff0000
            )
            embed.add_field(
                name="Validation Errors",
                value="\n".join(validation_errors[:10]),  # Limit to 10 errors
                inline=False
            )
            await ctx.send(embed=embed, delete_after=30)
            try:
                await ctx.message.delete()
            except discord.NotFound:
                pass
            return
        
        # Create backups and update files
        backup_info = []
        updated_files = []
        
        for target_path, file_content, original_filename in processed_files:
            try:
                # Create backup if original file exists
                if os.path.exists(target_path):
                    backup_path = f"{target_path}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                    with open(target_path, 'rb') as original:
                        with open(backup_path, 'wb') as backup:
                            backup.write(original.read())
                    backup_info.append(f"📋 Backup created: {os.path.basename(backup_path)}")
                
                # Write new file
                with open(target_path, 'wb') as f:
                    f.write(file_content)
                
                updated_files.append(original_filename)
                
            except Exception as e:
                validation_errors.append(f"❌ **{original_filename}** - Write error: {str(e)[:100]}")
                continue
        
        # Send success/failure report
        if updated_files:
            embed = discord.Embed(
                title="✅ Upload Successful",
                description=f"Successfully updated {len(updated_files)} TMB file(s)",
                color=0x00ff00,
                timestamp=datetime.now()
            )
            
            embed.add_field(
                name="📁 Updated Files",
                value="\n".join([f"✅ {filename}" for filename in updated_files]),
                inline=False
            )
            
            if backup_info:
                embed.add_field(
                    name="💾 Backups Created",
                    value="\n".join(backup_info),
                    inline=False
                )
            
            if validation_errors:
                embed.add_field(
                    name="⚠️ Validation Errors",
                    value="\n".join(validation_errors[:5]),  # Limit to 5 errors
                    inline=False
                )
            
            embed.add_field(
                name="👤 Uploaded by",
                value=ctx.author.mention,
                inline=True
            )
            
            embed.set_footer(text="Files have been validated and updated successfully")
            
            # Log the action
            print(f"📤 TMB files uploaded by {ctx.author} ({ctx.author.id}) in {ctx.guild.name}: {', '.join(updated_files)}")
        else:
            embed = discord.Embed(
                title="❌ Upload Failed",
                description="No files were updated due to validation errors.",
                color=0xff0000
            )
            embed.add_field(
                name="Validation Errors",
                value="\n".join(validation_errors[:10]),
                inline=False
            )
        
        await ctx.send(embed=embed, delete_after=60)
        
        # Delete the command message
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
            
    except Exception as e:
        await ctx.send(f"❌ Error processing TMB file uploads: {str(e)[:100]}...", delete_after=15)
        print(f"Error in upload_tmb_files command: {e}")

@bot.command(name='uploadarmory')
async def upload_armory_file(ctx):
    """
    Upload armory.json file to merge with existing armory data
    Usage: !uploadarmory (attach armory.json file)
    Restricted to Officers, Guild Leaders, and authorized roles
    """
    # Check permissions - only allow Officers, Guild Leaders
    authorized_roles = ["Officer", "Guild Leader"]
    user_roles = [role.name for role in ctx.author.roles]
    
    if not any(role in authorized_roles for role in user_roles):
        await ctx.send("❌ You don't have permission to use this command. Required roles: Officer or Guild Leader.", delete_after=10)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    # Get guild-specific file paths
    guild_id = ctx.guild.id
    paths = get_guild_file_paths(guild_id)
    
    # Initialize guild data if needed
    initialize_guild_data_files(guild_id)
    
    # Check if file is attached
    if not ctx.message.attachments:
        embed = discord.Embed(
            title="📤 Armory File Upload",
            description=f"Upload an armory.json file to merge with **{ctx.guild.name}** guild armory data.",
            color=0xff9900
        )
        embed.add_field(
            name="📋 File Requirements",
            value="• File must be named `armory.json`\n• Must contain valid JSON format\n• Data structure: `{\"PlayerName\": [\"Item1\", \"Item2\"]}`",
            inline=False
        )
        embed.add_field(
            name="📝 Instructions",
            value="1. Attach the `armory.json` file to your message\n2. Use the `!uploadarmory` command\n3. File will be validated and merged with existing data",
            inline=False
        )
        embed.add_field(
            name="⚠️ Merge Behavior",
            value="• New players will be added\n• New items for existing players will be added\n• Duplicate items will be ignored\n• Existing data will be preserved",
            inline=False
        )
        embed.add_field(
            name="💾 Backup",
            value="A timestamped backup of the current armory file will be created before merging.",
            inline=False
        )
        embed.add_field(
            name="🏰 Server Info",
            value=f"**Guild:** {ctx.guild.name}\n**File will be merged with this server's data only**",
            inline=False
        )
        embed.set_footer(text=f"Guild: {ctx.guild.name} • Data is server-specific")
        await ctx.send(embed=embed, delete_after=30)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    # Validate only one file
    if len(ctx.message.attachments) > 1:
        await ctx.send("❌ Please attach only one armory.json file.", delete_after=15)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    attachment = ctx.message.attachments[0]
    
    # Validate filename
    if attachment.filename.lower() != 'armory.json':
        await ctx.send("❌ File must be named `armory.json`. Please rename your file and try again.", delete_after=15)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    # Check file size (max 50MB for armory files)
    if attachment.size > 50 * 1024 * 1024:
        await ctx.send("❌ File too large (max 50MB). Please check your armory file.", delete_after=15)
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass
        return
    
    try:
        # Download and validate file content
        file_content = await attachment.read()
        
        try:
            # Parse JSON content
            uploaded_armory = json.loads(file_content.decode('utf-8'))
        except json.JSONDecodeError as e:
            embed = discord.Embed(
                title="❌ Invalid JSON Format",
                description="The uploaded file contains invalid JSON.",
                color=0xff0000
            )
            embed.add_field(
                name="Error Details",
                value=f"```{str(e)[:500]}```",
                inline=False
            )
            await ctx.send(embed=embed, delete_after=30)
            try:
                await ctx.message.delete()
            except discord.NotFound:
                pass
            return
        except UnicodeDecodeError:
            await ctx.send("❌ File encoding error. Please ensure the file is saved as UTF-8.", delete_after=15)
            try:
                await ctx.message.delete()
            except discord.NotFound:
                pass
            return
        
        # Validate armory data structure
        if not isinstance(uploaded_armory, dict):
            await ctx.send("❌ Invalid armory format. Expected JSON object with player names as keys.", delete_after=15)
            try:
                await ctx.message.delete()
            except discord.NotFound:
                pass
            return
        
        validation_errors = []
        valid_players = 0
        total_items = 0
        
        # Validate each player entry
        for player_name, items in uploaded_armory.items():
            if not isinstance(player_name, str) or not player_name.strip():
                validation_errors.append(f"Invalid player name: {repr(player_name)}")
                continue
            
            if not isinstance(items, list):
                validation_errors.append(f"Player '{player_name}': Items must be a list, got {type(items).__name__}")
                continue
            
            # Validate items
            for i, item in enumerate(items):
                if not isinstance(item, str):
                    validation_errors.append(f"Player '{player_name}', item {i+1}: Expected string, got {type(item).__name__}")
                    break
                if not item.strip():
                    validation_errors.append(f"Player '{player_name}', item {i+1}: Empty item name")
                    break
            else:
                # All items valid for this player
                valid_players += 1
                total_items += len(items)
        
        # Check if we have critical validation errors
        if validation_errors and valid_players == 0:
            embed = discord.Embed(
                title="❌ Validation Failed",
                description="The armory file contains critical errors and cannot be processed.",
                color=0xff0000
            )
            embed.add_field(
                name="Errors Found",
                value="\n".join(validation_errors[:10]) + ("..." if len(validation_errors) > 10 else ""),
                inline=False
            )
            await ctx.send(embed=embed, delete_after=30)
            try:
                await ctx.message.delete()
            except discord.NotFound:
                pass
            return
        
        # Load existing armory data using guild-specific paths
        existing_armory = {}
        if os.path.exists(paths['armory_file']):
            try:
                with open(paths['armory_file'], 'r', encoding='utf-8') as f:
                    existing_armory = json.load(f)
            except (json.JSONDecodeError, FileNotFoundError):
                existing_armory = {}
        
        # Create backup before merging using guild-specific paths
        backup_created = False
        backup_path = None
        if os.path.exists(paths['armory_file']):
            try:
                backup_path = f"{paths['armory_file']}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                with open(paths['armory_file'], 'rb') as original:
                    with open(backup_path, 'wb') as backup:
                        backup.write(original.read())
                backup_created = True
            except Exception as e:
                await ctx.send(f"❌ Failed to create backup: {str(e)[:100]}...", delete_after=15)
                try:
                    await ctx.message.delete()
                except discord.NotFound:
                    pass
                return
        
        # Merge armory data
        merge_stats = {
            'new_players': 0,
            'updated_players': 0,
            'new_items': 0,
            'duplicate_items': 0,
            'total_players_processed': 0
        }
        
        merged_armory = existing_armory.copy()
        
        for player_name, items in uploaded_armory.items():
            # Skip invalid entries
            if not isinstance(player_name, str) or not isinstance(items, list):
                continue
            
            clean_player_name = player_name.strip()
            if not clean_player_name:
                continue
            
            # Clean and validate items
            clean_items = []
            for item in items:
                if isinstance(item, str) and item.strip():
                    clean_items.append(item.strip())
            
            if not clean_items:
                continue
            
            merge_stats['total_players_processed'] += 1
            
            if clean_player_name not in merged_armory:
                # New player
                merged_armory[clean_player_name] = clean_items
                merge_stats['new_players'] += 1
                merge_stats['new_items'] += len(clean_items)
            else:
                # Existing player - merge items
                existing_items = set(merged_armory[clean_player_name])
                new_items = []
                
                for item in clean_items:
                    if item in existing_items:
                        merge_stats['duplicate_items'] += 1
                    else:
                        new_items.append(item)
                        merge_stats['new_items'] += 1
                
                if new_items:
                    merged_armory[clean_player_name].extend(new_items)
                    merge_stats['updated_players'] += 1
        
        # Save merged armory data using guild-specific paths
        try:
            # Sort merged armory data alphabetically by character name before saving
            sorted_merged_armory = dict(sorted(merged_armory.items()))
            
            # Write to temporary file first for atomic operation
            temp_file = paths['armory_file'] + '.tmp'
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(sorted_merged_armory, f, ensure_ascii=False, indent=2)
            
            # Atomic rename
            os.replace(temp_file, paths['armory_file'])
            
        except Exception as e:
            # Clean up temp file if it exists
            if os.path.exists(temp_file):
                os.remove(temp_file)
            await ctx.send(f"❌ Failed to save merged armory data: {str(e)[:100]}...", delete_after=15)
            try:
                await ctx.message.delete()
            except discord.NotFound:
                pass
            return
        
        # Create success report
        embed = discord.Embed(
            title="✅ Armory Upload & Merge Successful",
            description="Armory data has been successfully merged with existing data.",
            color=0x00ff00,
            timestamp=datetime.now()
        )
        
        # Merge statistics
        stats_text = []
        if merge_stats['new_players'] > 0:
            stats_text.append(f"👤 **{merge_stats['new_players']}** new players added")
        if merge_stats['updated_players'] > 0:
            stats_text.append(f"📝 **{merge_stats['updated_players']}** existing players updated")
        if merge_stats['new_items'] > 0:
            stats_text.append(f"⚔️ **{merge_stats['new_items']}** new items added")
        if merge_stats['duplicate_items'] > 0:
            stats_text.append(f"🔄 **{merge_stats['duplicate_items']}** duplicate items skipped")
        
        if stats_text:
            embed.add_field(
                name="📊 Merge Statistics",
                value="\n".join(stats_text),
                inline=False
            )
        
        # File info
        embed.add_field(
            name="📁 File Information",
            value=f"**Size:** {attachment.size:,} bytes\n**Players Processed:** {merge_stats['total_players_processed']}\n**Total Items in Upload:** {total_items}",
            inline=True
        )
        
        # Backup info
        if backup_created:
            embed.add_field(
                name="💾 Backup Created",
                value=f"`{os.path.basename(backup_path)}`",
                inline=True
            )
        
        # Validation warnings
        if validation_errors:
            embed.add_field(
                name="⚠️ Validation Warnings",
                value=f"{len(validation_errors)} entries skipped due to validation errors.\nProcessed {valid_players} valid players.",
                inline=False
            )
        
        embed.add_field(
            name="👤 Uploaded by",
            value=ctx.author.mention,
            inline=True
        )
        
        embed.set_footer(text="Armory data merged successfully • Use !get armory <player> to view player items")
        
        await ctx.send(embed=embed, delete_after=60)
        
        # Log the action
        print(f"📤 Armory file uploaded and merged by {ctx.author} ({ctx.author.id}) in {ctx.guild.name}")
        print(f"   Stats: {merge_stats['new_players']} new players, {merge_stats['new_items']} new items, {merge_stats['duplicate_items']} duplicates skipped")
        
    except Exception as e:
        await ctx.send(f"❌ Error processing armory upload: {str(e)[:100]}...", delete_after=15)
        print(f"Error in upload_armory_file command: {e}")
    finally:
        # Always delete the command message
        try:
            await ctx.message.delete()
        except discord.NotFound:
            pass

def get_character_role_sheet(player):
    """Return which role sheet a player belongs to: Tank, DPS, Caster, or Healers."""
    archetype = player.get("role", "") or ""
    sub_archetype = player.get("sub_archetype", "") or ""
    if archetype == "Tank":
        return "Tank"
    if archetype in ("Heal", "Healer"):
        return "Healers"
    if archetype == "DPS":
        return "Caster" if sub_archetype == "Caster" else "DPS"
    return None


def character_can_use_item(player, itemOffNotes):
    """
    Check if a player can use an item based on officer notes.

    Supports:
      New role tags : <Tank> <DPS> <Caster> <Healer>
      Class tags    : <Warrior> <Mage> etc.
      Spec+Class    : <ArcaneMage> <FuryWarrior> <RestoShaman> etc.
      Legacy role   : [Heal] [DPS] [Tank]  (backward compat with old TMB data)

    When both class tags and legacy [Role] brackets coexist the old combined
    logic is preserved so existing item-notes data keeps working correctly.
    """
    if not itemOffNotes:
        return False

    char_class     = player.get("class", "") or ""
    char_spec      = player.get("spec", "") or ""
    char_archetype = player.get("role", "") or ""
    char_sub       = player.get("sub_archetype", "") or ""

    all_tags      = re.findall(r'<([^>]+)>', itemOffNotes)
    old_role_tags = re.findall(r'\[([^\]]+)\]', itemOffNotes)

    if not all_tags and not old_role_tags:
        return False

    NEW_ROLE_TAGS = {"Tank", "DPS", "Caster", "Healer", "Healers", "Heal"}

    new_role_tags  = []
    class_tags     = []
    spec_class_tags = []  # list of (spec_prefix, class_name)

    for tag in all_tags:
        tag = tag.strip()
        if tag in NEW_ROLE_TAGS:
            new_role_tags.append(tag)
        elif tag in CLASS_LIST:
            class_tags.append(tag)
        else:
            matched = False
            for class_name in CLASS_LIST.keys():
                if len(tag) > len(class_name) and tag.lower().endswith(class_name.lower()):
                    spec_class_tags.append((tag[:-len(class_name)], class_name))
                    matched = True
                    break
            if not matched:
                class_tags.append(tag)

    # New role tags are plain OR conditions
    for tag in new_role_tags:
        if tag == "Tank" and char_archetype == "Tank":
            return True
        if tag == "DPS" and char_archetype == "DPS" and char_sub != "Caster":
            return True
        if tag == "Caster" and char_archetype == "DPS" and char_sub == "Caster":
            return True
        if tag in ("Healer", "Healers", "Heal") and char_archetype in ("Heal", "Healer"):
            return True

    # Spec+Class combined tags are also plain OR
    for spec_prefix, class_name in spec_class_tags:
        if char_class == class_name and char_spec.lower().startswith(spec_prefix.lower()):
            return True

    # Class tags with legacy [Role] brackets — preserve old combined logic
    if class_tags and old_role_tags:
        class_match = char_class in class_tags
        char_class_info = CLASS_LIST.get(char_class, {})
        char_class_roles = char_class_info.get("roles", [])

        if class_match:
            role_filter_applies = any(role in char_class_roles for role in old_role_tags)
            if not role_filter_applies:
                return True
            return char_archetype in old_role_tags
        else:
            for role in old_role_tags:
                role_matches = (char_archetype == role or
                                (role == "Heal" and char_archetype in ("Heal", "Healer")))
                if role_matches:
                    excluded = False
                    for listed_class in class_tags:
                        listed_info = CLASS_LIST.get(listed_class, {})
                        listed_roles = listed_info.get("roles", [])
                        if role in listed_roles or (role == "Heal" and "Heal" in listed_roles):
                            excluded = True
                            break
                    if not excluded:
                        return True
            return False

    elif class_tags:
        return char_class in class_tags

    elif old_role_tags:
        for role in old_role_tags:
            if role == "Heal" and char_archetype in ("Heal", "Healer"):
                return True
            if role == "Tank" and char_archetype == "Tank":
                return True
            if role == "DPS" and char_archetype == "DPS":
                return True

    return False


def createExcel(guild_id, excelType):
    """Create Excel file with guild-specific data"""
    # Get guild-specific file paths
    paths = get_guild_file_paths(guild_id)
    
    # Initialize guild data if needed
    initialize_guild_data_files(guild_id)
    players = {}
    wowClasses = []

    iconWidth = 4
    iconHeight = 25

    itemsIcons = {}
    playerParse = {}

    # Get Blizzard API token with retry logic
    access_token = None
    for attempt in range(3):
        try:
            response = requests.post(
                BLIZZARD_TOKEN_URL, 
                data={'grant_type': 'client_credentials'}, 
                auth=(BLIZZARD_ID, BLIZZARD_SECRET),
                timeout=10
            )
            response.raise_for_status()
            access_token = response.json()['access_token']
            break
        except requests.RequestException as e:
            print(f"⚠️ Token request attempt {attempt + 1} failed: {e}")
            if attempt == 2:
                print("❌ Failed to get Blizzard API token after 3 attempts")
                return

    with open(paths['item_icons_file'], 'r', encoding='utf-8') as f:
        try:
            itemsIcons = json.load(f)
            if not itemsIcons:
                itemsIcons = {}
        except json.JSONDecodeError:
            print("Warning: item-icons.json is corrupted or invalid. Starting with empty icons map.")
            itemsIcons = {}

    with open(paths['parses_file'], 'r', encoding='utf-8') as f:
        try:
            playerParse = json.load(f)
            if not playerParse:
                playerParse = {}
        except json.JSONDecodeError:
            print("Warning: parses.json is corrupted or invalid. Starting with empty parses data.")
            playerParse = {}

    #Attendance Start
    def get_reset_week_start(date_obj):
        # WoW reset week starts on Wednesday.
        days_since_wednesday = (date_obj.weekday() - 2) % 7
        return date_obj - timedelta(days=days_since_wednesday)

    def format_weekly_raid_note(note):
        note_clean = str(note or "").strip()
        note_lower = note_clean.lower()
        # For chill runs, do not expose split details in the weekly cell.
        if "chill run" in note_lower and "split" in note_lower:
            return "Chill run"
        return note_clean

    with open(paths['attendance_file'], newline='', encoding='utf-8') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=',', quotechar='"')
        
        firstRow = next(csvreader)
        attendanceDates = []
        attendanceWeekLabels = {}
        week_group_raid_events = {}
        
        for row in csvreader:
            raid_datetime = datetime.strptime(row[0].replace('"', ''), "%Y-%m-%d %H:%M:%S")
            week_start = get_reset_week_start(raid_datetime)
            week_key = week_start.strftime("%d/%m/%y")
            if week_key not in attendanceDates:
                attendanceDates.append(week_key)
                week_end = week_start + timedelta(days=6)
                attendanceWeekLabels[week_key] = f"{week_start.strftime('%d/%m/%y')} - {week_end.strftime('%d/%m/%y')}"
            
            playerName = row[2].replace('"', '').capitalize()
            raid_name = row[1].replace('"', '').strip() if len(row) > 1 else ""
            raid_note = row[11].replace('"', '').strip() if len(row) > 11 else ""
            raid_group = row[13].replace('"', '').strip() if len(row) > 13 else ""
            player = {}
            raids = []
            benchedRaids = []
            absentRaids = []
            unpreparedRaids = []
            raidNotesByWeek = {}
            raidGroupsByWeek = {}

            if week_key not in week_group_raid_events:
                week_group_raid_events[week_key] = {}
            if raid_group not in week_group_raid_events[week_key]:
                week_group_raid_events[week_key][raid_group] = set()

            # Count unique raid events for the group in this reset week.
            # Keying by (raid_datetime, raid_name) collapses per-character rows
            # while still counting distinct raids of the same group in that week.
            raid_datetime_key = raid_datetime.strftime("%Y-%m-%d %H:%M:%S")
            raid_event_key = (raid_datetime_key, raid_name)
            week_group_raid_events[week_key][raid_group].add(raid_event_key)

            try: 
                if players[playerName]:
                    player = players[playerName]
                    raids = player["raids"]
                    benchedRaids = player["benched_raids"]
                    absentRaids = player["absent_raids"]
                    unpreparedRaids = player["unprepared_raids"]
                    raidNotesByWeek = player.get("raid_notes_by_week", {})
                    raidGroupsByWeek = player.get("raid_groups_by_week", {})
            except:
                player["name"] = playerName

            if week_key not in raidNotesByWeek:
                raidNotesByWeek[week_key] = []
            if raid_note and raid_note not in raidNotesByWeek[week_key]:
                raidNotesByWeek[week_key].append(raid_note)

            if week_key not in raidGroupsByWeek:
                raidGroupsByWeek[week_key] = []
            if raid_group and raid_group not in raidGroupsByWeek[week_key]:
                raidGroupsByWeek[week_key].append(raid_group)
            
            if row[6].replace('"', '') == "Benched":
                if week_key not in benchedRaids:
                    benchedRaids.append(week_key)
            elif row[6].replace('"', '') == "Gave notice":
                if week_key not in absentRaids:
                    absentRaids.append(week_key)
            elif row[6].replace('"', '') == "Unprepared":
                if week_key not in unpreparedRaids:
                    unpreparedRaids.append(week_key)
            else:
                if week_key not in raids:
                    raids.append(week_key)
            
            player["raids"] = raids
            player["benched_raids"] = benchedRaids
            player["absent_raids"] = absentRaids
            player["unprepared_raids"] = unpreparedRaids
            player["raid_notes_by_week"] = raidNotesByWeek
            player["raid_groups_by_week"] = raidGroupsByWeek
            player["firstRaid"] = week_key
            player["isInAttendance"] = True
            players[playerName] = player

        attendanceDates.sort(key=lambda d: datetime.strptime(d, "%d/%m/%y"))
    #Attendance Finish
    
    #Loot Start
    playerData = ""
    with open(paths['character_file'], 'r', encoding='utf-8') as file:
        playerData = json.load(file)
    
    for playerInfo in playerData :
        player = {}
        name = ""
        try:
            name = playerInfo["name"].capitalize()
            player = players[name]
        except:
            player["name"] = playerInfo["name"].capitalize()
            player["role"] = playerInfo["archetype"]
            player["firstRaid"] = "31/12/30"
            player["raids"] = []
            player["benched_raids"] = []
            player["absent_raids"] = []
            player["unprepared_raids"] = []
            player["raid_notes_by_week"] = {}
            player["raid_groups_by_week"] = {}
            player["wishlist"] = "0/0"
            print("Player " + name + " is not present in the attendance file. Probably a new player or it's unclaimed on TMB.")
        
        player["isInLootData"] = True
        loot = {}
        
        for lootReceived in playerInfo["received"]:
            lootItem = {}
            lootItem["name"] = lootReceived["name"]
            lootItem["id"] = lootReceived["item_id"]
            lootItem["isOS"] = lootReceived["pivot"]["is_offspec"]
            
            date_object = datetime.strptime(lootReceived["pivot"]["received_at"], "%Y-%m-%d %H:%M:%S")
            formatted_date = date_object.strftime("%d/%m/%y")
            lootItem["receivedDate"] = formatted_date
            loot[lootItem["name"]] = lootItem
            
        wishlist = []
        sum = 0
        try:
            wishlist = playerInfo["wishlist"]
        except:
            print("No wishlist found for " + name)
        for wishlistItem in playerInfo["wishlist"]:
            if wishlistItem["pivot"]["is_received"] == 1:
                sum += 1
        
        if playerInfo["class"] is not None and not playerInfo["class"] in wowClasses:
            wowClasses.append(playerInfo["class"])

        playerParseData = playerParse.get(name, {})
        player["bestPerformanceAverage"] = playerParseData.get("bestPerformanceAverage", 0.0)
        player["medianPerformanceAverage"] = playerParseData.get("medianPerformanceAverage", 0.0)

        player["race"] = playerInfo["race"]
        player["role"] = playerInfo["archetype"]
        player["spec"] = playerInfo.get("spec") or "-"
        player["class"] = playerInfo["class"]
        player["is_alt"] = playerInfo["is_alt"]
        player["sub_archetype"] = playerInfo.get("sub_archetype") or ""
        player["member_id"] = playerInfo["member_id"]
        player["raid_group_name"] = playerInfo.get("raid_group_name") or "Unknown"
        player["raid_group_color"] = playerInfo.get("raid_group_color")
        player["wishlist"] = str(sum)+"/"+str(len(wishlist))
        player["loot"] = loot
        discord_name = (playerInfo.get("username") or "-").strip()
        if discord_name and discord_name != "-":
            discord_name = discord_name[:1].upper() + discord_name[1:]
        player["discord_username"] = discord_name
        
        players[name] = player
    #Loot Finish
    
    #First Sheet Start
    # Wishlist is intentionally hidden for now (easy to restore later).
    # To re-enable: add "Wishlist" back after "Last MS" and restore the commented Wishlist append block below.
    # To restore Last MS: insert "Last MS" after "MS Ratio"
    # To restore Last bench: insert "Last bench" after "Parses (Best|Median)"
    column_names = ["Character", "Class", "Spec", "Race", "Player", "Raid", "Char", "Raids", "Benched", "Attendance", "Items (+OS)", "MS Ratio", "Parses (Best|Median)", "Character"]
    counter = 0
    for date in attendanceDates :
        counter += 1
        if counter == 20:
            counter = 0
            column_names.append("Name")
        column_names.append(attendanceWeekLabels.get(date, date))   
    
    playerInfoList = []
    raid_group_colors = {}
    player_roles = {}
    
    for player in list(players.values()):
        try:
            if not player["isInLootData"]:
                print("Removed player " + player["name"] + " since he's not in the roster.")
                del players[player["name"]]
                continue
        except:
            print("Removed player " + player["name"] + " since he's not in the roster.")
            del players[player["name"]]
            continue
            
        playerInfo = []
        playerDateInfo = []
        totalRaids = 0
        completedRaids = 0
        benchedRaids = 0
        absentRaids = 0
        attendance = 0.0
        itemsPlaceholders = "<MS>(+<OS>)"
        lastReceivedItemDate = "-"
        lastBench = "-"
        msItems = 0
        msRatio = 0.0
        osItems = 0
        
        counter = 0
        for date in attendanceDates :

            counter += 1
            if counter == 20:
                counter = 0
                try:
                    name = player["name"].capitalize()
                    if player["member_id"] is None :
                        name += " (Unclaimed)"
                    playerDateInfo.append(name)
                except:
                    playerDateInfo.append("-")

            found = False
            benched = False
            unprepared_b = False
            
            for raid in player["raids"]:
                if date == raid :
                    completedRaids += 1
                    found = True
                
            for bench in player["benched_raids"]:
                if date == bench :
                    benchedRaids += 1
                    benched = True
                    
            for unprepared in player["unprepared_raids"]:
                if date == unprepared :
                    unprepared_b = True
                
            week_notes = player.get("raid_notes_by_week", {}).get(date, [])
            week_groups = player.get("raid_groups_by_week", {}).get(date, [])

            show_week_note = False
            for group_name in week_groups:
                group_raid_events = week_group_raid_events.get(date, {}).get(group_name, set())
                # Multiple raids means at least 2 distinct raid events for the same group.
                if len(group_raid_events) > 1:
                    show_week_note = True
                    break

            display_notes = []
            for week_note in week_notes:
                formatted_note = format_weekly_raid_note(week_note)
                if formatted_note and formatted_note not in display_notes:
                    display_notes.append(formatted_note)
            note_text = ", ".join(display_notes) if display_notes and show_week_note else ""

            def build_week_cell(value_text):
                return (note_text + " | " + value_text) if note_text else value_text

            if not found:
                if benched:
                    if lastBench == "-":
                        lastBench = date
                    playerDateInfo.append(build_week_cell("Benched"))
                elif unprepared_b:
                    playerDateInfo.append(build_week_cell("Holiday"))
                else:
                    if datetime.strptime(player["firstRaid"], "%d/%m/%y") > datetime.strptime(date, "%d/%m/%y"):
                        playerDateInfo.append(build_week_cell("N/A"))
                    else:
                        absentRaids += 1
                        playerDateInfo.append(build_week_cell("Absent"))
            else:
                currentMsItems = 0
                currentOsItems = 0
                
                lootReceived = {}
                try:
                    lootReceived = player["loot"]
                except:
                    lootReceived = {}
                    
                for loot in lootReceived.values():
                    loot_date_obj = datetime.strptime(loot["receivedDate"], "%d/%m/%y")
                    loot_week_key = get_reset_week_start(loot_date_obj).strftime("%d/%m/%y")
                    if loot_week_key == date:
                        if loot["isOS"] == 0:
                            if lastReceivedItemDate == "-":
                                lastReceivedItemDate = date
                            currentMsItems += 1
                        else:
                            currentOsItems += 1
                if currentMsItems == 0 and currentOsItems == 0:
                    appendStr = "-"
                else:
                    appendStr = itemsPlaceholders.replace("<MS>", str(currentMsItems)).replace("<OS>", str(currentOsItems))
                playerDateInfo.append(build_week_cell(appendStr))
                
                msItems += currentMsItems
                osItems += currentOsItems
        
        if completedRaids > 0:
            msRatio = f"{msItems / completedRaids:.2f}"
            osRatio = f"{osItems / completedRaids:.2f}"
        
        totalRaids = completedRaids + benchedRaids + absentRaids

        #Name
        try:
            name = player["name"].capitalize()
            if player["member_id"] is None :
                name += " (Unclaimed)"
            playerInfo.append(name)
        except:
            playerInfo.append("-")
        #Class
        try:
            playerInfo.append(player["class"])
        except:
            playerInfo.append("-")
        #Spec
        try:
            playerInfo.append(player.get("spec", "-"))
        except:
            playerInfo.append("-")
        #Race
        try:
            playerInfo.append(player["race"])
        except:
            playerInfo.append("-")
        #Discord
        try:
            playerInfo.append(player.get("discord_username", "-"))
        except:
            playerInfo.append("-")
        #Split
        try:
            playerInfo.append(player.get("raid_group_name", "Unknown"))
        except:
            playerInfo.append("Unknown")
        #Char (Main/Alt)
        playerInfo.append("Alt" if player.get("is_alt") else "Main")
        #Completed Raids
        playerInfo.append(completedRaids)
        #Benched Raids
        playerInfo.append(benchedRaids)
        #Attendance
        if totalRaids > 0:
            attendance = ((completedRaids + benchedRaids) / totalRaids)
        player["attendance"] = f"{attendance*100:.1f}" 
        playerInfo.append(attendance)
        #MS Items
        playerInfo.append(str(msItems) + " (" + str(osItems) + ")")
        playerInfo.append(msRatio)
        player["msRatio"] = msRatio
        #Last MS (disabled - to restore: add "Last MS" after "MS Ratio" in column_names)
        # playerInfo.append(lastReceivedItemDate)
        #Wishlist (disabled for now)
        # try:
        #     playerInfo.append(player["wishlist"])
        # except:
        #     playerInfo.append("0/0")
        #Parse Info (merged as Best|Median)
        best_avg = player.get("bestPerformanceAverage", 0)
        median_avg = player.get("medianPerformanceAverage", 0)
        
        # Ensure values are numeric, default to 0 if not
        try:
            best_avg = float(best_avg) if best_avg not in [None, "", "-"] else 0
        except (ValueError, TypeError):
            best_avg = 0
            
        try:
            median_avg = float(median_avg) if median_avg not in [None, "", "-"] else 0
        except (ValueError, TypeError):
            median_avg = 0
        
        best_str = f"{best_avg:.1f}" if best_avg != 0 else "-"
        median_str = f"{median_avg:.1f}" if median_avg != 0 else "-"
        playerInfo.append(f"{best_str} | {median_str}")
        #Last bench (disabled - to restore: add "Last bench" after "Parses (Best|Median)" in column_names)
        # playerInfo.append(lastBench)
        #Spacer
        try:
            name = player["name"].capitalize()
            if player["member_id"] is None :
                name += " (Unclaimed)"
            playerInfo.append(name)
        except:
            playerInfo.append("-")
        
        #Raids Presence and Loot
        for dateInfo in playerDateInfo:
            playerInfo.append(dateInfo)
        
        #Add
        playerInfoList.append(playerInfo)
        player_roles[playerInfo[0]] = str(player.get("role") or "").strip().upper()
        raid_group_name = player.get("raid_group_name") or "Unknown"
        raid_group_color = player.get("raid_group_color")
        if raid_group_name not in raid_group_colors and raid_group_color:
            raid_group_colors[raid_group_name] = raid_group_color
    
    workbook = Workbook()
    sheet = workbook.active

    def render_attendance_sheet(target_sheet, sheet_name, attendance_rows):
        target_sheet.title = sheet_name[:31]
        target_sheet.auto_filter.ref = "A1:BZ1"

        def normalize_raid_group_color(raw_color):
            if raw_color is None:
                return None

            color_str = str(raw_color).strip()
            if not color_str:
                return None

            if color_str.startswith("#"):
                color_str = color_str[1:]

            if len(color_str) == 6:
                try:
                    int(color_str, 16)
                    return color_str.upper()
                except ValueError:
                    pass

            try:
                return f"{int(float(color_str)):06X}"[-6:]
            except (ValueError, TypeError):
                return None

        def soften_hex_color(hex_color, blend_ratio=0.72):
            """Blend a hex color toward white to improve readability in sheet cells."""
            if not hex_color or len(hex_color) != 6:
                return hex_color
            try:
                r = int(hex_color[0:2], 16)
                g = int(hex_color[2:4], 16)
                b = int(hex_color[4:6], 16)
            except ValueError:
                return hex_color

            r = int(r + (255 - r) * blend_ratio)
            g = int(g + (255 - g) * blend_ratio)
            b = int(b + (255 - b) * blend_ratio)
            return f"{r:02X}{g:02X}{b:02X}"

        # Sort by Class, then Attendance.
        sorted_data = sorted(attendance_rows, key=lambda x: (x[1], x[9]), reverse=True)

        for col_num, column_name in enumerate(column_names, start=1):
            cell = target_sheet.cell(row=1, column=col_num, value=column_name)
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center")

        data_row_count = len(sorted_data)
        for row_num, row_data in enumerate(sorted_data, start=2):
            for col_num, cell_value in enumerate(row_data, start=1):
                cell = target_sheet.cell(row=row_num, column=col_num, value=cell_value)

                cell.alignment = Alignment(horizontal="center")
                if col_num == 8 or col_num == 9 or col_num == 12:
                    for row in range(2, data_row_count + 2):
                        target_sheet.cell(row=row, column=col_num).number_format = "0"  # Numeric format
                if col_num == 10:
                    for row in range(2, data_row_count + 2):
                        target_sheet.cell(row=row, column=col_num).number_format = "0.00%"
                # DD/MM date format - disabled (Last MS was col 12, Last bench was col 14)
                # if col_num == 12 or col_num == 14:
                #     for row in range(2, data_row_count + 2):
                #         target_sheet.cell(row=row, column=col_num).number_format = "DD/MM"

        column_sizes = {
            "Character": 22,
            "Class": 12,
            "Spec": 17,
            "Race": 12,
            "Raids": 10,
            "Benched": 13,
            "Attendance": 16,
            "Items (+OS)": 14,
            "MS Ratio": 12,
            "Last MS": 14,
            "Parses (Best|Median)": 24,
            "Char": 10,
            "OS Items": 16,
            "OS Ratio": 14,
            "Last Bench": 16,
            "Player": 22,
            "Raid": 12,
            "Default": 22
        }
        # Adjust column widths
        for col_num, column_name in enumerate(column_names, start=1):
            column_letter = get_column_letter(col_num)

            try:
                size = column_sizes[column_name]
            except:
                size = column_sizes["Default"]

            target_sheet.column_dimensions[column_letter].width = max(len(column_name), size)

            for row in range(2, 80):
                cell = target_sheet[column_letter + str(row)]
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="center")

                    thin_border = Side(border_style="thin", color="000000")  # Black thin border
                    cell.border = Border(top=thin_border, bottom=thin_border, left=thin_border, right=thin_border)

                    cell.font = Font(name="Aptos Light", bold=False)

                    if column_name == "Character" or column_name == "Class":
                        value = cell.value
                        if column_name == "Character":
                            cell.alignment = Alignment(horizontal="left")
                            cell.font = Font(name="Aptos", bold=True)
                            value = target_sheet["B" + str(row)].value
                        bgcolor = CLASS_LIST.get(value, {}).get("color", "CCCCCC")
                        cell.fill = PatternFill(start_color=bgcolor, end_color=bgcolor, fill_type="solid")
                    elif column_name == "Spec":
                        cell.alignment = Alignment(horizontal="left")
                        cell.font = Font(name="Aptos", bold=True, color="1F2937")
                        role = player_roles.get(target_sheet["A" + str(row)].value, "")
                        role_color = {
                            "DPS": "E8B4B8",
                            "HEAL": "B7E4C7",
                            "TANK": "B6CCFE",
                        }.get(role)
                        if role_color:
                            cell.fill = PatternFill(start_color=role_color, end_color=role_color, fill_type="solid")
                    elif column_name == "Player":
                        cell.alignment = Alignment(horizontal="left")
                        cell.font = Font(name="Aptos", bold=True)
                    elif column_name == "Raid":
                        cell.alignment = Alignment(horizontal="left")
                        cell.font = Font(name="Aptos", bold=True, color="1F2937")
                        fill_color = normalize_raid_group_color(raid_group_colors.get(str(cell.value)))
                        if fill_color:
                            soft_fill_color = soften_hex_color(fill_color)
                            cell.fill = PatternFill(start_color=soft_fill_color, end_color=soft_fill_color, fill_type="solid")
                    elif column_name == "Char":
                        cell.alignment = Alignment(horizontal="left")
                        cell.font = Font(name="Aptos", bold=True, color="1F2937")
                        if cell.value == "Main":
                            cell.fill = PatternFill(start_color="CFEAD6", end_color="CFEAD6", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="D6E4FF", end_color="D6E4FF", fill_type="solid")
                    elif column_name == "Race":
                        bgcolor = "CCCCCC"
                        if cell.value == "Dwarf":
                            bgcolor = "C69B6D"
                        elif cell.value == "Gnome":
                            bgcolor = "FFF468"
                        elif cell.value == "Human":
                            bgcolor = "F48CBA"
                        elif cell.value == "Draenei":
                            bgcolor = "00FF98"
                        elif cell.value == "Blood Elf":
                            bgcolor = "33937F"
                        elif cell.value == "Night Elf":
                            bgcolor = "0070DD"
                        elif cell.value == "Orc":
                            bgcolor = "AAD372"
                        elif cell.value == "Tauren":
                            bgcolor = "C69B6D"
                        elif cell.value == "Troll":
                            bgcolor = "FFFFFF"
                        elif cell.value == "Undead":
                            bgcolor = "8788EE"
                        cell.fill = PatternFill(start_color=bgcolor, end_color=bgcolor, fill_type="solid")
                    elif column_name == "Attendance":
                        start_color = (255, 156, 0)
                        end_color = (117, 249, 77)
                        percentage = cell.value
                        color = calculate_gradient_color(percentage, start_color, end_color)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    elif column_name == "MS Ratio":
                        start_color = (255, 255, 255)
                        end_color = (186, 72, 177)
                        if cell.value != "-" and cell.value is not None:
                            percentage = float(cell.value)
                            color = calculate_gradient_color(percentage, start_color, end_color)
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    elif column_name == "Parses (Best|Median)":
                        if cell.value and cell.value != "- | -":
                            try:
                                best_part = str(cell.value).split("|")[0].strip()
                                value = float(best_part) if best_part != "-" else 0
                            except (ValueError, TypeError):
                                value = 0

                            if value > 0:
                                color = "66664E"
                                if value == 100:
                                    color = "E5A93F"
                                elif value >= 99:
                                    color = "BE49A8"
                                elif value >= 95:
                                    color = "FF8000"
                                elif value >= 75:
                                    color = "A335EE"
                                elif value >= 50:
                                    color = "0961FE"
                                elif value >= 25:
                                    color = "0961FE"
                                cell.font = Font(name="Aptos", bold=True, color=color)
                    elif col_num > 14:
                        bgcolor = "CCCCCC"
                        cell.font = Font(name="Aptos Light", bold=True)
                        cell_value = str(cell.value)
                        base_value = cell_value
                        raid_note = ""
                        if " | " in cell_value:
                            raid_note, base_value = cell_value.rsplit(" | ", 1)
                            base_value = base_value.strip()
                            raid_note = raid_note.strip().lower()

                        if base_value == "N/A":
                            bgcolor = "FFFFFF"
                            cell.value = ""
                        elif base_value == "Benched":
                            bgcolor = "9DC0FA"
                        elif base_value == "Absent":
                            bgcolor = "FF9C00"
                        elif base_value == "Holiday":
                            bgcolor = "00FFFF"
                        elif base_value == "-":
                            if "Split 1" in raid_note:
                                if raid_group_name == "Speed Run":
                                    bgcolor = "8BEFA8"
                                else:
                                    bgcolor = "FFD98A"
                            elif "Split 2" in raid_note:
                                bgcolor = "7FD1FF"
                            else:
                                bgcolor = "A1FB8E"
                        else:
                            if "Split 1" in raid_note:
                                if raid_group_name == "Speed Run":
                                    bgcolor = "72E89A"
                                else:
                                    bgcolor = "F7CA6A"
                            elif "Split 2" in raid_note:
                                bgcolor = "69C4F7"
                            else:
                                bgcolor = "75F94D"
                            if base_value.startswith("0"):
                                cell.font = Font(name="Aptos Light", bold=False)
                        cell.fill = PatternFill(start_color=bgcolor, end_color=bgcolor, fill_type="solid")

        print("Create attendance sheet '" + target_sheet.title + "' with " + str(data_row_count) + " players.")
    
    if excelType == "Attendance" or excelType == "All" :
        render_attendance_sheet(sheet, "Attendance", playerInfoList)
    else:
        # If we're not creating an attendance sheet, remove the empty active sheet
        # We'll add other sheets later, so this prevents an empty first sheet
        workbook.remove(sheet)
    #First Sheet Stop

    # Item Sheets Start
    # Loading file
    itemList = {}

    with open(paths['item_file'], newline='', encoding='utf-8') as csvfile:
        csvreader = csv.reader(csvfile, delimiter=',', quotechar='"')
        
        firstRow = next(csvreader)
        
        for row in csvreader:
            itemName = row[0].replace('"', '')
            itemId = row[1]
            itemInstance = row[2].replace('"', '')
            itemSource = row[3].replace('"', '')

            itemNotes = row[5].replace('"', '')
            itemOffNotes = row[6].replace('"', '')
            tier_label = row[8].replace('"', '')

            item = {}
            item["itemName"] = itemName
            item["itemId"] = itemId
            item["itemInstance"] = itemInstance
            item["itemSource"] = itemSource
            item["itemNotes"] = itemNotes
            item["itemOffNotes"] = itemOffNotes
            item["tier_label"] = tier_label

            if tier_label and itemOffNotes:
                itemList[itemId] = item
    # Loading file finish

    #Load armory cache
    with open(paths['armory_file'], 'r', encoding='utf-8') as f:
        try:
            armoryList = json.load(f)
            if not armoryList:
                armoryList = {}
        except json.JSONDecodeError:
            print("Warning: armory.json is corrupted or invalid. Starting with empty armory list.")
            armoryList = {}


    if excelType == "Loot" or excelType == "All" :
        raidList = {
            "Molten Core": "E26B0A",
            "Blackwing Lair": "C0504D",
            "Temple of Ahn'Qiraj": "4F6228",
            "Naxxramas": "403151",
            "Magtheridon's Lair": "403151",
            "Gruul's Lair": "C0504D",
            "Tempest Keep": "DAB1DA",
            "Serpentshrine Cavern": "4F6228",
        }

        loot_raid_groups = sorted(set(p.get("raid_group_name", "Unknown") for p in players.values()))
        if not loot_raid_groups:
            loot_raid_groups = ["Unknown"]

        for loot_raid_group in loot_raid_groups:
            loot_group_players = {name: p for name, p in players.items()
                                  if p.get("raid_group_name") == loot_raid_group}
            if not loot_group_players:
                continue

            sheet_title = f"{loot_raid_group} Loot"
            if len(sheet_title) > 31:
                sheet_title = sheet_title[:31]
            allLootSheet = workbook.create_sheet(title=sheet_title)

            i = 1
            for raid in raidList.keys():
                raidItems = {}
                for itemId, item in itemList.items():
                    if item["itemInstance"] == raid:
                        raidItems[itemId] = item
                if not raidItems:
                    continue
                allLootSheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
                cell = allLootSheet.cell(row=i, column=1, value=raid)
                cell.fill = PatternFill(start_color=raidList[raid], end_color=raidList[raid], fill_type="solid")
                i += 1
                for raidItem in raidItems.values():
                    raidItemName = raidItem["itemName"]
                    itemOffNotes = raidItem["itemOffNotes"]
                    lootRow = ["", raidItem["itemId"], raidItemName, raidItem["itemNotes"], raidItem["tier_label"], "", "", ""]
                    for player in loot_group_players.values():
                        if not character_can_use_item(player, itemOffNotes):
                            continue
                        playerName = player["name"].capitalize()
                        found = False
                        try:
                            playerArmory = armoryList[playerName]
                        except KeyError:
                            playerArmory = []
                            armoryList[playerName] = playerArmory
                        for armoryItem in armoryList[playerName]:
                            if armoryItem == raidItemName:
                                found = True
                            elif raidItemName == "Head of Nefarian":
                                if armoryItem == "Master Dragonslayer's Medallion" or armoryItem == "Master Dragonslayer's Orb" or armoryItem == "Master Dragonslayer's Ring":
                                    found = True
                            elif raidItemName == "Eye of C'Thun":
                                if armoryItem == "Amulet of the Fallen God" or armoryItem == "Cloak of the Fallen God" or armoryItem == "Ring of the Fallen God":
                                    found = True
                            elif raidItemName == "Vek'nilash's Circlet":
                                if armoryItem == "Conqueror's Crown" or armoryItem == "Doomcaller's Circlet" or armoryItem == "Enigma Circlet" or armoryItem == "Tiara of the Oracle":
                                    found = True
                        for loot in player["loot"].values():
                            if loot["name"] == raidItemName:
                                found = True
                        if not found:
                            alt_label = " [A]" if player.get("is_alt") else ""
                            lootRow.append(f'{playerName}{alt_label} ({player["attendance"]}% - {player["msRatio"]})')
                    allLootSheet.append(lootRow)
                    i += 1

            instanceColor = raidList.get("Molten Core", "E26B0A")
            thin = Side(border_style="thin", color="000000")
            for row in allLootSheet.iter_rows(min_row=1, max_row=allLootSheet.max_row):
                if row[0].value is None or row[0].value == "" and row[1].value is None or row[1].value == "":
                    allLootSheet.row_dimensions[row[0].row].height = iconHeight
                    continue

                if row[0].value is not None and row[0].value != "":
                    allLootSheet.row_dimensions[row[0].row].height = iconHeight
                    row[0].font = Font(name="Aptos", bold=True, color="FFFFFF", size=16)
                    row[0].alignment = Alignment(horizontal="center", vertical="center")
                    instanceColor = row[0].fill.start_color.index
                    continue

                row[0].fill = PatternFill(start_color=instanceColor, end_color=instanceColor, fill_type="solid")
                item_id_cell = row[1]
                item_id = str(item_id_cell.value)
                allLootSheet.row_dimensions[item_id_cell.row].height = iconHeight

                if item_id not in itemsIcons.keys():
                    try:
                        media_url = f'https://eu.api.blizzard.com/data/wow/media/item/{item_id}?namespace=static-classic-eu&locale=en_GB'
                        urlHeaders = {'Authorization': f'Bearer {access_token}'}
                        media_response = requests.get(media_url, headers=urlHeaders)
                        icon_url = media_response.json()['assets'][0]['value']
                        itemsIcons[item_id] = icon_url
                    except:
                        print("Error fetching media for loot item:", raidItemName)

                icon_url = itemsIcons.get(item_id)
                if icon_url:
                    item_id_cell.value = f'=IMAGE("{icon_url}", 2)'
                row[1].border = Border(left=thin, right=thin, top=thin, bottom=thin)

                row[2].alignment = Alignment(horizontal="left", vertical="center")
                row[2].font = Font(name="Aptos", bold=True)
                row[2].border = Border(left=thin, right=thin, top=thin, bottom=thin)

                if row[3].value is not None and row[3].value != "":
                    notes = row[3].value
                    row[3].value = '=IMAGE("https://render.worldofwarcraft.com/classic-eu/icons/56/inv_misc_questionmark.jpg", 2)'
                    row[3].comment = Comment(text=notes, author="")
                else:
                    allLootSheet.merge_cells(start_row=item_id_cell.row, start_column=3, end_row=item_id_cell.row, end_column=4)
                row[3].border = Border(left=thin, right=thin, top=thin, bottom=thin)

                row[4].alignment = Alignment(horizontal="center", vertical="center")
                row[4].font = Font(name="Aptos", bold=True)
                row[4].border = Border(left=thin, right=thin, top=thin, bottom=thin)

                if row[4].value == "1" or row[4].value == "S":
                    row[4].fill = PatternFill(start_color="32C3F6", end_color="32C3F6", fill_type="solid")
                elif row[4].value == "2" or row[4].value == "A":
                    row[4].fill = PatternFill(start_color="20FF26", end_color="20FF26", fill_type="solid")
                elif row[4].value == "3" or row[4].value == "B":
                    row[4].fill = PatternFill(start_color="F7FF26", end_color="F7FF26", fill_type="solid")
                elif row[4].value == "4" or row[4].value == "C":
                    row[4].fill = PatternFill(start_color="FF734D", end_color="FF734D", fill_type="solid")
                elif row[4].value == "5" or row[4].value == "D":
                    row[4].fill = PatternFill(start_color="F30026", end_color="F30026", fill_type="solid")
                elif row[4].value == "6" or row[4].value == "F":
                    row[4].fill = PatternFill(start_color="CC3071", end_color="CC3071", fill_type="solid")

                foundTheEnd = False
                index = 8
                while not foundTheEnd:
                    if index >= len(row):
                        foundTheEnd = True
                        break
                    cell = row[index]
                    if cell.value is None or cell.value == "":
                        foundTheEnd = True
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = Font(name="Aptos", bold=True)
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                        raw_name = cell.value.split(" [A]")[0].split(" (")[0].strip()
                        player = loot_group_players.get(raw_name.capitalize())
                        if player:
                            classColor = CLASS_LIST.get(player["class"], {}).get("color", "CCCCCC")
                            cell.fill = PatternFill(start_color=classColor, end_color=classColor, fill_type="solid")

                    index += 1
                    if index >= len(row):
                        foundTheEnd = True

            for column in allLootSheet.columns:
                column_letter = get_column_letter(column[0].column)
                column_size = 30
                if column_letter == "B" or column_letter == "D" or column_letter == "E":
                    column_size = 4.5
                elif column_letter == "A":
                    column_size = 4
                elif column_letter == "C":
                    column_size = 45
                allLootSheet.column_dimensions[column_letter].width = column_size

    if excelType == "Class Items" or excelType == "All" :
        ROLE_SHEET_DEFS = [
            {"key": "Tank",    "label": "Tanks",   "color": "7B6941"},
            {"key": "DPS",     "label": "DPS",     "color": "9B2335"},
            {"key": "Caster",  "label": "Casters", "color": "2255A4"},
            {"key": "Healers", "label": "Healers", "color": "2E7D32"},
        ]

        role_raid_groups = sorted(set(p.get("raid_group_name", "Unknown") for p in players.values()))
        if not role_raid_groups:
            role_raid_groups = ["Unknown"]

        # Pre-compute items for each role using all players across all raid groups
        role_items_cache = {}
        for role_def in ROLE_SHEET_DEFS:
            role_key = role_def["key"]
            role_players_global = [p for p in players.values() if get_character_role_sheet(p) == role_key]
            role_items_cache[role_key] = [
                item for item in itemList.values()
                if any(character_can_use_item(p, item["itemOffNotes"]) for p in role_players_global)
            ]
            print(f"Role {role_key} has {len(role_items_cache[role_key])} items.")

        for role_def in ROLE_SHEET_DEFS:
            role_key   = role_def["key"]
            role_label = role_def["label"]
            role_color = role_def["color"]
            role_items = role_items_cache[role_key]

            if not role_items:
                continue

            for rg_name in role_raid_groups:
                sheet_players = {
                    name: p for name, p in players.items()
                    if p.get("raid_group_name") == rg_name and get_character_role_sheet(p) == role_key
                }
                if not sheet_players:
                    continue

                sheet_title = f"{rg_name} {role_label}"
                if len(sheet_title) > 31:
                    sheet_title = sheet_title[:31]
                roleSheet = workbook.create_sheet(title=sheet_title)

                headers = [" ", " ", " ", " ", " "]
                for player in sheet_players.values():
                    headers.append(player["name"].capitalize())
                headers.append(" ")

                for col_num, header in enumerate(headers, start=1):
                    thin = Side(border_style="thin", color="000000")
                    column_letter = get_column_letter(col_num)
                    cell = roleSheet.cell(row=1, column=col_num, value=header)
                    cell.fill = PatternFill(start_color=role_color, end_color=role_color, fill_type="solid")
                    if cell.value != " ":
                        cell.border = Border(thin, thin, thin, thin)
                    cell.font = Font(name="Aptos", bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    column_size = 16
                    if col_num == 1:
                        column_size = 4
                    elif col_num == 2:
                        column_size = 4.5
                    elif col_num == 3:
                        column_size = 45
                    elif col_num == 4:
                        column_size = 5
                    elif col_num == 5:
                        column_size = 5
                    elif col_num == len(sheet_players) + 6:
                        column_size = 4
                    roleSheet.column_dimensions[column_letter].width = column_size

                totalRows = len(role_items) + 2
                for item in role_items:
                    itemData = ["", item["itemId"], item["itemName"], item["itemNotes"], item["tier_label"]]
                    roleSheet.append(itemData)

                for row in roleSheet.iter_rows(min_row=2, max_row=roleSheet.max_row):
                    row[0].fill = PatternFill(start_color=role_color, end_color=role_color, fill_type="solid")
                    row[len(sheet_players) + 5].fill = PatternFill(start_color=role_color, end_color=role_color, fill_type="solid")
                    item_id_cell = row[1]
                    item_id = str(item_id_cell.value)

                    roleSheet.row_dimensions[item_id_cell.row].height = iconHeight

                    if item_id is None or item_id == "":
                        current_row = item_id_cell.row
                        roleSheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=len(sheet_players) + 5)
                        row[2].alignment = Alignment(horizontal="left", vertical="top")
                        row[2].font = Font(name="Aptos", bold=False)
                        row[2].fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
                        continue

                    current_item = next((it for it in role_items if it["itemId"] == item_id), None)

                    if item_id not in itemsIcons.keys():
                        try:
                            media_url = f'https://eu.api.blizzard.com/data/wow/media/item/{item_id}?namespace=static-classic-eu&locale=en_GB'
                            urlHeaders = {'Authorization': f'Bearer {access_token}'}
                            media_response = requests.get(media_url, headers=urlHeaders)
                            icon_url = media_response.json()['assets'][0]['value']
                            itemsIcons[item_id] = icon_url
                        except:
                            print(f"Error fetching media for item {item_id}")

                    icon_url = itemsIcons.get(item_id)
                    if icon_url:
                        item_id_cell.value = f'=IMAGE("{icon_url}", 2)'

                    row[2].alignment = Alignment(horizontal="left", vertical="center")
                    row[2].font = Font(name="Aptos", bold=True)

                    if row[3].value is not None and row[3].value != "":
                        notes = row[3].value
                        row[3].value = '=IMAGE("https://render.worldofwarcraft.com/classic-eu/icons/56/inv_misc_questionmark.jpg", 2)'
                        row[3].comment = Comment(text=notes, author="")
                    else:
                        roleSheet.merge_cells(start_row=item_id_cell.row, start_column=3, end_row=item_id_cell.row, end_column=4)

                    row[4].alignment = Alignment(horizontal="center", vertical="center")
                    row[4].font = Font(name="Aptos", bold=True)

                    if row[4].value == "1" or row[4].value == "S":
                        row[4].fill = PatternFill(start_color="32C3F6", end_color="32C3F6", fill_type="solid")
                    elif row[4].value == "2" or row[4].value == "A":
                        row[4].fill = PatternFill(start_color="20FF26", end_color="20FF26", fill_type="solid")
                    elif row[4].value == "3" or row[4].value == "B":
                        row[4].fill = PatternFill(start_color="F7FF26", end_color="F7FF26", fill_type="solid")
                    elif row[4].value == "4" or row[4].value == "C":
                        row[4].fill = PatternFill(start_color="FF734D", end_color="FF734D", fill_type="solid")
                    elif row[4].value == "5" or row[4].value == "D":
                        row[4].fill = PatternFill(start_color="F30026", end_color="F30026", fill_type="solid")
                    elif row[4].value == "6" or row[4].value == "F":
                        row[4].fill = PatternFill(start_color="CC3071", end_color="CC3071", fill_type="solid")

                    actual_item_name = current_item["itemName"] if current_item else ""

                    for col_idx, (pname, playerInfo) in enumerate(sheet_players.items(), start=5):
                        currCell = row[col_idx]
                        currCell.alignment = Alignment(horizontal="center", vertical="center")
                        currCell.value = "-"

                        try:
                            playerArmory = armoryList[pname]
                        except KeyError:
                            playerArmory = []
                            armoryList[pname] = playerArmory

                        for armoryItem in armoryList[pname]:
                            found = False
                            if armoryItem == actual_item_name:
                                found = True
                            elif actual_item_name == "Head of Nefarian":
                                if armoryItem in ("Master Dragonslayer's Medallion", "Master Dragonslayer's Orb", "Master Dragonslayer's Ring"):
                                    found = True
                            if found:
                                currCell.value = "Equipped"
                                currCell.fill = PatternFill(start_color="A1FB8E", end_color="A1FB8E", fill_type="solid")
                                break

                        for loot in playerInfo["loot"].values():
                            if loot["name"] == actual_item_name:
                                alt_tag = " [A]" if playerInfo.get("is_alt") else ""
                                currCell.value = f"LC {loot['receivedDate']}{alt_tag}"
                                currCell.fill = PatternFill(start_color="75F94D", end_color="75F94D", fill_type="solid")
                                break

                        if current_item and not character_can_use_item(playerInfo, current_item["itemOffNotes"]):
                            if currCell.value == "-":
                                currCell.value = "OS"
                                currCell.fill = PatternFill(start_color="9DC0FA", end_color="9DC0FA", fill_type="solid")

                rows_to_remove = []
                for row_idx in range(2, roleSheet.max_row + 1):
                    item_id_check = roleSheet.cell(row=row_idx, column=2)
                    if item_id_check.value is None or item_id_check.value == "":
                        continue
                    all_os_or_empty = True
                    has_any_value = False
                    for col_num in range(6, len(sheet_players) + 6):
                        cell_value = roleSheet.cell(row=row_idx, column=col_num).value
                        if cell_value and cell_value != "-":
                            has_any_value = True
                            if cell_value != "OS":
                                all_os_or_empty = False
                                break
                    if has_any_value and all_os_or_empty:
                        rows_to_remove.append(row_idx)
                        item_name_check = roleSheet.cell(row=row_idx, column=3).value
                        print(f"Removing '{item_name_check}' from {sheet_title} — all players OS")

                for row_idx in reversed(rows_to_remove):
                    roleSheet.delete_rows(row_idx, 1)
                if rows_to_remove:
                    totalRows -= len(rows_to_remove)

                for col_num in range(1, len(sheet_players) + 7):
                    finalCell = roleSheet.cell(row=totalRows, column=col_num)
                    finalCell.fill = PatternFill(start_color=role_color, end_color=role_color, fill_type="solid")

                min_row = 1
                max_row = roleSheet.max_row
                min_col = 1
                max_col = roleSheet.max_column
                thick = Side(border_style="thick", color="000000")
                thin  = Side(border_style="thin",  color="000000")

                for row_num in range(min_row, max_row + 1):
                    for col_num in range(min_col, max_col + 1):
                        cell = roleSheet.cell(row=row_num, column=col_num)
                        left_border   = col_num == min_col + 1 and row_num > min_row and row_num < max_row
                        right_border  = col_num == max_col - 1 and row_num > min_row and row_num < max_row
                        top_border    = row_num == min_row + 1 and col_num > min_col and col_num < max_col
                        bottom_border = row_num == max_row - 1 and col_num > min_col and col_num < max_col
                        if row_num > min_row and row_num < max_row and col_num > min_col and col_num < max_col:
                            cell.border = Border(thin, thin, thin, thin)
                        b = cell.border
                        border = Border(
                            left=thick   if (col_num == min_col or left_border)   else b.left,
                            right=thick  if (col_num == max_col or right_border)  else b.right,
                            top=thick    if (row_num == min_row or top_border)    else b.top,
                            bottom=thick if (row_num == max_row or bottom_border) else b.bottom,
                        )
                        cell.border = border

    # Item Sheets Finish

    #Save cache item icons
    with open(paths['item_icons_file'], 'w', encoding='utf-8') as f:
        json.dump(itemsIcons, f, ensure_ascii=False, indent=4)

    # Ensure we have at least one sheet in the workbook
    if len(workbook.worksheets) == 0:
        # Create a minimal info sheet if no other sheets were created
        info_sheet = workbook.create_sheet(title="Info")
        info_sheet.cell(row=1, column=1, value="No data available for the requested report type.")
        info_sheet.cell(row=2, column=1, value=f"Report type: {excelType}")
        info_sheet.cell(row=3, column=1, value="Please check your data files and try again.")

    # Return the workbook for sending to Discord
    return workbook

# Add better error handling for bot startup
if __name__ == "__main__":
    try:
        print("🤖 Starting HopiumBot...")
        logger.info("Starting HopiumBot...")
        print(f"Token present: {'Yes' if token else 'No'}")
        logger.info(f"Token present: {'Yes' if token else 'No'}")
        
        # Use INFO level for production, DEBUG for development
        log_level = logging.INFO if os.getenv('RENDER') else logging.DEBUG
        logger.info(f"Environment: {'Render (Production)' if os.getenv('RENDER') else 'Local Development'}")
        logger.info(f"Log level: {log_level}")
        
        # Start HTTP server for Render health checks (only in production)
        if os.getenv('RENDER'):
            async def health_check(request):
                return web.Response(text="HopiumBot is running!")
            
            async def start_web_server():
                app = web.Application()
                app.router.add_get('/', health_check)
                app.router.add_get('/health', health_check)
                
                port = int(os.environ.get('PORT', 8080))
                runner = web.AppRunner(app)
                await runner.setup()
                site = web.TCPSite(runner, '0.0.0.0', port)
                await site.start()
                print(f"🌐 Health check server started on port {port}")
                logger.info(f"Health check server started on port {port}")
            
            # Start web server in background
            async def main():
                await start_web_server()
                await bot.start(token)
            
            asyncio.run(main())
        else:
            # Local development - no web server needed
            logger.info("🚀 Starting bot in local development mode")
            bot.run(token)
            
    except discord.LoginFailure:
        error_msg = "Invalid bot token. Please check your DISCORD_TOKEN environment variable."
        print(f"❌ ERROR: {error_msg}")
        logger.error(error_msg)
        print("1. Go to https://discord.com/developers/applications")
        print("2. Select your application > Bot")
        print("3. Reset Token and update your environment variables")
    except discord.HTTPException as e:
        if "PHONE_REGISTRATION_ERROR" in str(e):
            error_msg = "PHONE_REGISTRATION_ERROR: This is a Discord account/token issue."
            print(f"❌ {error_msg}")
            logger.error(error_msg)
            print("Solutions:")
            print("1. Regenerate your bot token")
            print("2. Check if your Discord account needs phone verification")
            print("3. Wait 24-48 hours and try again")
        else:
            print(f"❌ HTTP Error: {e}")
            logger.error(f"HTTP Error: {e}")
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        logger.error(f"Unexpected error: {e}", exc_info=True)
        # Exit gracefully in production
        import sys
        sys.exit(1)